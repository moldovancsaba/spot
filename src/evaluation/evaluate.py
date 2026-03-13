from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from ..pipeline import run_classification


def _read_assigned_category_by_row(path: Path, max_items: int | None = None) -> Dict[int, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    cat_idx = header.index("Assigned Category") + 1

    try:
        _min_col, _min_row, _max_col, max_row = range_boundaries(ws.calculate_dimension())
    except Exception:
        max_row = ws.max_row

    result: Dict[int, str] = {}
    for row_idx in range(2, max_row + 1):
        v = ws.cell(row=row_idx, column=cat_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            result[row_idx] = s
            if max_items is not None and len(result) >= max_items:
                break
    return result


def _read_flags_by_row(path: Path, row_filter: set[int] | None = None) -> Dict[int, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    flags_idx = header.index("Flags") + 1

    result: Dict[int, str] = {}
    if row_filter is not None:
        for row_idx in sorted(row_filter):
            v = ws.cell(row=row_idx, column=flags_idx).value
            result[row_idx] = "" if v is None else str(v)
        return result

    try:
        _min_col, _min_row, _max_col, max_row = range_boundaries(ws.calculate_dimension())
    except Exception:
        max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        v = ws.cell(row=row_idx, column=flags_idx).value
        result[row_idx] = "" if v is None else str(v)
    return result


def _read_integrity(run_dir: Path) -> dict:
    p = run_dir / "integrity_report.json"
    if not p.exists():
        raise RuntimeError(f"Missing integrity report: {p}")
    return json.loads(p.read_text(encoding="utf-8"))


def _ensure_run_ok(run_dir: Path) -> None:
    progress = json.loads((run_dir / "progress.json").read_text(encoding="utf-8"))
    if progress.get("state") != "COMPLETED":
        raise RuntimeError(f"Run failed or incomplete: {run_dir.name}, state={progress.get('state')}")
    integrity = _read_integrity(run_dir)
    if not integrity.get("canonical_set_validation_passed", False):
        raise RuntimeError(f"Integrity failure in run: {run_dir.name}")


def evaluate_runs(
    input_path: Path,
    ssot_path: Path,
    runs_dir: Path,
    evaluation_run_id: str,
    run_language: str,
    review_mode: str,
    single_model: str,
    ensemble_models: list[str],
    max_workers: int = 1,
    limit: int | None = None,
    progress_every: int = 100,
) -> Path:
    eval_dir = runs_dir / evaluation_run_id
    eval_dir.mkdir(parents=True, exist_ok=True)

    single_run_id = f"{evaluation_run_id}-single"
    ensemble_run_id = f"{evaluation_run_id}-ensemble"

    single_output = eval_dir / "single_output.xlsx"
    ensemble_output = eval_dir / "ensemble_output.xlsx"

    single_run_dir = runs_dir / single_run_id
    ensemble_run_dir = runs_dir / ensemble_run_id

    if not (single_run_dir / "progress.json").exists() or not single_output.exists():
        run_classification(
            input_path=input_path,
            output_path=single_output,
            run_id=single_run_id,
            run_language=run_language,
            review_mode=review_mode,
            ssot_path=ssot_path,
            runs_dir=runs_dir,
            max_workers=max_workers,
            limit=limit,
            ensemble_enabled=False,
            ensemble_models=[single_model],
            consensus_strategy="majority",
            disagreement_mode="human_review",
            progress_every=progress_every,
        )

    if not (ensemble_run_dir / "progress.json").exists() or not ensemble_output.exists():
        run_classification(
            input_path=input_path,
            output_path=ensemble_output,
            run_id=ensemble_run_id,
            run_language=run_language,
            review_mode=review_mode,
            ssot_path=ssot_path,
            runs_dir=runs_dir,
            max_workers=max_workers,
            limit=limit,
            ensemble_enabled=True,
            ensemble_models=ensemble_models,
            consensus_strategy="majority",
            disagreement_mode="human_review",
            progress_every=progress_every,
        )

    _ensure_run_ok(single_run_dir)
    _ensure_run_ok(ensemble_run_dir)

    single_integrity = _read_integrity(single_run_dir)
    expected_processed = int(single_integrity.get("total_processed", 0))

    single_map = _read_assigned_category_by_row(single_output, max_items=expected_processed if expected_processed > 0 else None)
    ensemble_map = _read_assigned_category_by_row(ensemble_output, max_items=expected_processed if expected_processed > 0 else None)

    common_rows = sorted(set(single_map.keys()) & set(ensemble_map.keys()))
    ensemble_flags = _read_flags_by_row(ensemble_output, row_filter=set(common_rows))
    total_rows = len(common_rows)
    if total_rows == 0:
        raise RuntimeError("No comparable rows found between single and ensemble outputs")

    identical = 0
    changed = 0
    change_matrix = defaultdict(lambda: defaultdict(int))
    single_dist: Counter = Counter()
    ensemble_dist: Counter = Counter()

    consensus_high = 0
    consensus_medium = 0
    consensus_low = 0
    disagreement_count = 0

    for row in common_rows:
        s = single_map[row]
        e = ensemble_map[row]
        single_dist[s] += 1
        ensemble_dist[e] += 1
        if s == e:
            identical += 1
        else:
            changed += 1
            change_matrix[s][e] += 1

        flags = ensemble_flags.get(row, "")
        if "CONSENSUS_HIGH" in flags:
            consensus_high += 1
        elif "CONSENSUS_MEDIUM" in flags:
            consensus_medium += 1
        elif "CONSENSUS_LOW" in flags:
            consensus_low += 1
        if "DISAGREEMENT" in flags:
            disagreement_count += 1

    categories = sorted(set(single_dist.keys()) | set(ensemble_dist.keys()))
    shift = {c: ensemble_dist.get(c, 0) - single_dist.get(c, 0) for c in categories}

    report = {
        "evaluation_run_id": evaluation_run_id,
        "single_run_id": single_run_id,
        "ensemble_run_id": ensemble_run_id,
        "total_rows": total_rows,
        "identical_classifications_count": identical,
        "identical_percentage": round((identical / total_rows) * 100, 4),
        "changed_classifications_count": changed,
        "changed_percentage": round((changed / total_rows) * 100, 4),
        "consensus_high_count": consensus_high,
        "consensus_medium_count": consensus_medium,
        "consensus_low_count": consensus_low,
        "disagreement_count": disagreement_count,
        "single_model_distribution": dict(single_dist),
        "ensemble_distribution": dict(ensemble_dist),
        "distribution_shift_per_category": shift,
        "change_matrix": {k: dict(v) for k, v in change_matrix.items()},
    }

    (eval_dir / "single_run_id").write_text(single_run_id, encoding="utf-8")
    (eval_dir / "ensemble_run_id").write_text(ensemble_run_id, encoding="utf-8")
    out = eval_dir / "evaluation_report.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return out
