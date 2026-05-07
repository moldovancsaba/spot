from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import List, Set

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries

from .models import ClassificationResult, InputRow, SSOT


class InputFileError(RuntimeError):
    pass


MAX_INPUT_FILE_BYTES = 25 * 1024 * 1024
MAX_INPUT_ROWS = 100000
MAX_POST_TEXT_LENGTH = 20000
ORIGINAL_ROW_INDEX_COLUMN = "Original Row Index"


ADDED_COLUMNS = [
    "Raw Category",
    "Assigned Category",
    "Consensus Tier",
    "Minority Report",
    "Model Votes",
    "Fallback Events",
    "Judge Score",
    "Judge Verdict",
    "Confidence Score",
    "Soft Signal Score",
    "Soft Signal Flags",
    "Soft Signal Evidence",
    "Explanation / Reasoning",
    "Flags",
    "Model Version",
    "Prompt Version",
    "Taxonomy Version",
    "SSOT Version",
    "Pipeline Version",
    "Run ID",
    "Run Language",
    "Review Mode",
    "Review Required",
    "Row Hash",
]


def read_input_rows(path: Path, ssot: SSOT) -> List[InputRow]:
    if path.suffix.lower() != ".xlsx":
        raise InputFileError("Input must be .xlsx")
    if path.stat().st_size > MAX_INPUT_FILE_BYTES:
        raise InputFileError(
            f"Input workbook is too large. Maximum supported size is {MAX_INPUT_FILE_BYTES // (1024 * 1024)} MiB."
        )

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InputFileError(
            f"Corrupted/invalid workbook: {exc}. Expected a valid .xlsx with columns: {ssot.policy.expected_columns}"
        ) from exc

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise InputFileError("Input worksheet is empty. Expected header row.")

    header_values = [str(v).strip() if v is not None else "" for v in header]
    expected = ssot.policy.expected_columns
    if header_values[: len(expected)] != expected:
        raise InputFileError(
            f"Invalid schema. Expected first columns {expected}; got {header_values[:len(expected)]}. Fix input schema to match the sample format."
        )

    try:
        _min_col, _min_row, _max_col, max_row = range_boundaries(ws.calculate_dimension())
    except Exception:
        max_row = ws.max_row

    original_row_index_idx = header_values.index(ORIGINAL_ROW_INDEX_COLUMN) if ORIGINAL_ROW_INDEX_COLUMN in header_values else None

    rows: List[InputRow] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row, values_only=True), start=2):
        item_number = "" if row[0] is None else str(row[0]).strip()
        post_text = "" if row[1] is None else str(row[1]).strip()
        category = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
        if item_number == "" and post_text == "" and category == "":
            continue
        if len(post_text) > MAX_POST_TEXT_LENGTH:
            raise InputFileError(
                f"Input row {idx} exceeds the maximum supported post text length of {MAX_POST_TEXT_LENGTH} characters."
            )
        if "\x00" in post_text:
            raise InputFileError(f"Input row {idx} contains invalid null-byte content.")
        original_row_index = idx
        if original_row_index_idx is not None and original_row_index_idx < len(row):
            raw_original_row_index = row[original_row_index_idx]
            if raw_original_row_index not in {None, ""}:
                try:
                    original_row_index = int(raw_original_row_index)
                except Exception as exc:  # noqa: BLE001
                    raise InputFileError(f"Input row {idx} contains an invalid Original Row Index value.") from exc
        rows.append(InputRow(row_index=original_row_index, item_number=item_number, post_text=post_text))
        if len(rows) > MAX_INPUT_ROWS:
            raise InputFileError(f"Input workbook exceeds the maximum supported row count of {MAX_INPUT_ROWS}.")

    if not rows:
        raise InputFileError("Input worksheet has no data rows.")

    return rows


def stable_row_hash(item_number: str, post_text: str) -> str:
    payload = json.dumps({"item_number": item_number, "post_text": post_text}, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_row_manifest_entries(rows: List[InputRow]) -> List[dict]:
    entries: List[dict] = []
    for sequence_index, row in enumerate(rows, start=1):
        text = row.post_text or ""
        entries.append(
            {
                "sequence_index": sequence_index,
                "row_index": int(row.row_index),
                "item_number": row.item_number,
                "row_hash": stable_row_hash(row.item_number, text),
                "post_text_sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
                "post_text_length": len(text),
            }
        )
    return entries


def build_stored_row_entries(rows: List[InputRow]) -> List[dict]:
    entries: List[dict] = []
    for sequence_index, row in enumerate(rows, start=1):
        text = row.post_text or ""
        entries.append(
            {
                "sequence_index": sequence_index,
                "row_index": int(row.row_index),
                "item_number": row.item_number,
                "post_text": text,
                "source_category": "",
                "row_hash": stable_row_hash(row.item_number, text),
                "post_text_sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
                "post_text_length": len(text),
            }
        )
    return entries


def write_row_manifest(path: Path, entries: List[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for entry in entries:
            handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


def build_segment_input_workbook_from_entries(output_path: Path, entries: List[dict]) -> List[dict]:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="Sheet")
    worksheet.append(["Item number", "Post text", "Category", ORIGINAL_ROW_INDEX_COLUMN])
    manifest_entries: List[dict] = []
    for entry in entries:
        item_number = str(entry.get("item_number") or "")
        post_text_value = str(entry.get("post_text") or "")
        source_category = entry.get("source_category")
        row_index = int(entry["row_index"])
        worksheet.append([item_number, post_text_value, source_category or None, row_index])
        manifest_entries.append(
            {
                "sequence_index": int(entry["sequence_index"]),
                "row_index": row_index,
                "item_number": item_number,
                "row_hash": str(entry["row_hash"]),
                "post_text_sha256": str(entry["post_text_sha256"]),
                "post_text_length": int(entry["post_text_length"]),
            }
        )
    workbook.save(output_path)
    return manifest_entries


def ensure_output_columns(output_path: Path) -> None:
    wb = load_workbook(output_path)
    try:
        ws = wb[wb.sheetnames[0]]
        header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if all(column in header for column in ADDED_COLUMNS):
            return
        start_col = ws.max_column + 1
        for offset, col_name in enumerate(ADDED_COLUMNS):
            ws.cell(row=1, column=start_col + offset, value=col_name)
        wb.save(output_path)
    finally:
        wb.close()


def build_segment_input_workbook(input_path: Path, output_path: Path, row_start: int, row_end: int) -> List[dict]:
    source = load_workbook(input_path, read_only=True, data_only=True)
    manifest_entries: List[dict] = []
    try:
        source_ws = source[source.sheetnames[0]]
        header = [value for value in next(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title=source_ws.title)
        worksheet.append([*header[:3], ORIGINAL_ROW_INDEX_COLUMN])
        data_row_index = 0
        for excel_row_index, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
            item_number = "" if row[0] is None else str(row[0]).strip()
            post_text = "" if row[1] is None else str(row[1]).strip()
            category = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
            if item_number == "" and post_text == "" and category == "":
                continue
            data_row_index += 1
            if data_row_index < row_start or data_row_index > row_end:
                continue
            worksheet.append([row[0], row[1], row[2] if len(row) >= 3 else None, excel_row_index])
            item_number = "" if row[0] is None else str(row[0]).strip()
            post_text_value = "" if row[1] is None else str(row[1]).strip()
            manifest_entries.append(
                {
                    "sequence_index": data_row_index,
                    "row_index": excel_row_index,
                    "item_number": item_number,
                    "row_hash": stable_row_hash(item_number, post_text_value),
                    "post_text_sha256": hashlib.sha256(post_text_value.encode("utf-8")).hexdigest(),
                    "post_text_length": len(post_text_value),
                }
            )
        workbook.save(output_path)
    finally:
        source.close()
    return manifest_entries


def merge_segment_output(segment_output_path: Path, aggregate_output_path: Path) -> None:
    ensure_output_columns(aggregate_output_path)
    source = load_workbook(segment_output_path, read_only=True, data_only=True)
    target = load_workbook(aggregate_output_path)
    try:
        source_ws = source[source.sheetnames[0]]
        target_ws = target[target.sheetnames[0]]
        source_header = [str(v).strip() if v is not None else "" for v in next(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        target_header = [str(v).strip() if v is not None else "" for v in next(target_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if ORIGINAL_ROW_INDEX_COLUMN not in source_header:
            raise InputFileError("Segment output is missing Original Row Index and cannot be merged.")
        source_original_row_idx = source_header.index(ORIGINAL_ROW_INDEX_COLUMN)
        source_output_indices = {name: source_header.index(name) for name in ADDED_COLUMNS if name in source_header}
        target_output_columns = {name: target_header.index(name) + 1 for name in ADDED_COLUMNS if name in target_header}
        missing_columns = [name for name in ADDED_COLUMNS if name not in target_output_columns]
        if missing_columns:
            raise InputFileError(f"Aggregate output is missing required output columns: {missing_columns}")
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            raw_original_row = row[source_original_row_idx] if source_original_row_idx < len(row) else None
            if raw_original_row in {None, ""}:
                continue
            target_row = int(raw_original_row)
            for column_name, source_idx in source_output_indices.items():
                value = row[source_idx] if source_idx < len(row) else None
                target_ws.cell(row=target_row, column=target_output_columns[column_name], value=value)
        target.save(aggregate_output_path)
    finally:
        source.close()
        target.close()


def write_output(
    input_path: Path,
    output_path: Path,
    ssot: SSOT,
    run_id: str,
    run_language: str,
    review_mode: str,
    pipeline_version: str,
    results: List[ClassificationResult],
    row_hashes: List[str],
) -> None:
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    start_col = ws.max_column + 1
    for offset, col_name in enumerate(ADDED_COLUMNS):
        ws.cell(row=1, column=start_col + offset, value=col_name)

    for i, result in enumerate(results, start=2):
        if not result.category:
            result.category = "Not Antisemitic"
            result.flags = sorted(set(result.flags + ["EMPTY_CATEGORY_RECOVERED"]))
        review_required = "YES" if "REVIEW_REQUIRED" in result.flags else "NO"
        values = [
            result.raw_category,
            result.category,
            result.consensus_tier,
            result.minority_label,
            "" if not result.model_votes else ";".join(f"{k}:{v}" for k, v in sorted(result.model_votes.items())),
            "" if not result.fallback_events else ";".join(sorted(set(result.fallback_events))),
            result.judge_score,
            result.judge_verdict,
            round(result.confidence, 4),
            "" if result.soft_signal_score is None else round(result.soft_signal_score, 4),
            "" if not result.soft_signal_flags else ";".join(sorted(set(result.soft_signal_flags))),
            "" if not result.soft_signal_evidence else " | ".join(result.soft_signal_evidence),
            result.explanation,
            ";".join(sorted(set(result.flags))),
            result.resolved_model_version or ssot.policy.model_version,
            ssot.policy.prompt_version,
            ssot.taxonomy.version,
            ssot.ssot_version,
            pipeline_version,
            run_id,
            run_language,
            review_mode,
            review_required,
            row_hashes[i - 2],
        ]
        write_row = result.row_index if result.row_index > 1 else i
        for offset, value in enumerate(values):
            ws.cell(row=write_row, column=start_col + offset, value=value)

    wb.save(output_path)


def validate_no_null_assigned_category(output_path: Path, expected_rows: List[int]) -> None:
    wb = load_workbook(output_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        category_col_idx = header.index("Assigned Category") + 1
    except ValueError as exc:
        raise InputFileError("Output validation failed: 'Assigned Category' column missing") from exc

    null_rows: List[int] = []
    for row_idx in expected_rows:
        val = ws.cell(row=row_idx, column=category_col_idx).value
        if val is None or str(val).strip() == "":
            null_rows.append(row_idx)

    if null_rows:
        sample = null_rows[:10]
        raise InputFileError(
            f"Output validation failed: null Assigned Category in {len(null_rows)} rows. Sample rows: {sample}"
        )


def extract_assigned_categories(output_path: Path, expected_rows: List[int]) -> Set[str]:
    wb = load_workbook(output_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        category_col_idx = header.index("Assigned Category") + 1
    except ValueError as exc:
        raise InputFileError("Output validation failed: 'Assigned Category' column missing") from exc

    values: Set[str] = set()
    for row_idx in expected_rows:
        val = ws.cell(row=row_idx, column=category_col_idx).value
        if val is None:
            continue
        txt = str(val).strip()
        if txt:
            values.add(txt)
    return values
