from __future__ import annotations

import json
import time
from pathlib import Path

from backend.services.ops_db_service import (
    build_run_segment_summary,
    count_other_runs_for_upload,
    delete_run_snapshot,
    delete_upload_snapshot,
    fetch_run_row,
    fetch_run_rows,
    fetch_upload_rows_by_row_indices,
    latest_run_attempt,
    register_run,
    summarize_run_rows,
    update_latest_run_attempt_status,
    update_run_row_review,
    update_run_snapshot,
    upsert_run_rows,
)
from backend.services.artifact_manifest_service import RUN_ARTIFACT_NAMES
from backend.services.excel_service import UPLOADS_DIRNAME, read_upload_record
from backend.services.run_lifecycle_service import (
    discover_run_process_pid,
    normalize_state_from_segments,
)
from openpyxl import load_workbook
from src.excel_io import ORIGINAL_ROW_INDEX_COLUMN


RUN_RECORD = "run_record.json"
REVIEW_STATE = "review_state.json"
SIGNOFF = "signoff.json"
ACTION_LOG = "action_log.jsonl"
ARTIFACT_NAMES = RUN_ARTIFACT_NAMES
RUN_HISTORY_DIRNAME = "_history"


def run_dir(runs_dir: Path, run_id: str) -> Path:
    return runs_dir / run_id


def run_history_dir(runs_dir: Path) -> Path:
    path = runs_dir / RUN_HISTORY_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def run_record_path(runs_dir: Path, run_id: str) -> Path:
    return run_dir(runs_dir, run_id) / RUN_RECORD


def review_state_path(runs_dir: Path, run_id: str) -> Path:
    return run_dir(runs_dir, run_id) / REVIEW_STATE


def signoff_path(runs_dir: Path, run_id: str) -> Path:
    return run_dir(runs_dir, run_id) / SIGNOFF


def action_log_path(runs_dir: Path, run_id: str) -> Path:
    return run_dir(runs_dir, run_id) / ACTION_LOG


def safe_read_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _fallback_processing_stats(record: dict, progress: dict | list | None) -> dict:
    progress_dict = progress if isinstance(progress, dict) else {}
    processed_rows = int(progress_dict.get("processed_rows") or 0)
    total_rows = int(progress_dict.get("total_rows") or 0)
    created_at = int(record.get("created_at") or 0)
    updated_at = int(record.get("updated_at") or created_at or 0)
    elapsed_seconds = max(updated_at - created_at, 0)
    return {
        "processed_rows": processed_rows,
        "total_rows": total_rows,
        "elapsed_seconds": elapsed_seconds,
        "avg_seconds_per_row": round(elapsed_seconds / processed_rows, 4) if processed_rows > 0 else None,
        "threat_rows_detected": None,
        "threat_rate": None,
        "projected_threat_rows": None,
        "review_required_rows_detected": None,
        "judged_rows": 0,
        "second_pass_candidates": None,
        "second_pass_completed": None,
        "second_pass_overrides": None,
    }


def _resolved_total_rows(*, record: dict, progress: dict | None, segment_summary: dict | None) -> int:
    segment_total_rows = int((segment_summary or {}).get("total_rows") or 0)
    if segment_total_rows > 0:
        return segment_total_rows
    if isinstance(progress, dict) and progress.get("total_rows") not in {None, ""}:
        return int(progress.get("total_rows"))
    return int(record.get("total_rows") or 0)


def _resolved_processed_rows(*, record: dict, progress: dict | None, canonical_stats: dict | None, segment_summary: dict | None) -> int:
    canonical_processed_rows = int((canonical_stats or {}).get("processed_rows") or 0)
    if canonical_processed_rows > 0:
        return canonical_processed_rows
    segment_processed_rows = int((segment_summary or {}).get("processed_rows") or 0)
    if segment_processed_rows > 0:
        return segment_processed_rows
    if isinstance(progress, dict):
        return int(progress.get("processed_rows") or 0)
    return int(record.get("processed_rows") or 0)


def append_action(
    *,
    runs_dir: Path,
    run_id: str,
    action: str,
    actor: str = "local-operator",
    payload: dict | None = None,
) -> None:
    path = action_log_path(runs_dir, run_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    entry = {
        "timestamp": int(time.time()),
        "actor": actor,
        "action": action,
        "payload": payload or {},
    }
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


def read_action_log(*, runs_dir: Path, run_id: str) -> list[dict]:
    path = action_log_path(runs_dir, run_id)
    if not path.exists():
        return []
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except Exception:
            continue
    return rows


def create_run_record(
    *,
    runs_dir: Path,
    run_id: str,
    input_path: str,
    output_path: str,
    upload_id: str | None,
    language: str,
    review_mode: str,
    start_payload: dict | None = None,
) -> dict:
    record = {
        "run_id": run_id,
        "upload_id": upload_id,
        "input_path": input_path,
        "output_path": output_path,
        "language": language,
        "review_mode": review_mode,
        "start_payload": start_payload or {},
        "state": "STARTING",
        "review_summary": {
            "review_required_rows": 0,
            "reviewed_rows": 0,
            "pending_rows": 0,
        },
        "created_at": int(time.time()),
        "updated_at": int(time.time()),
    }
    write_run_record(runs_dir=runs_dir, run_id=run_id, record=record)
    register_run(
        runs_dir=runs_dir,
        run_id=run_id,
        upload_id=upload_id,
        language=language,
        review_mode=review_mode,
        state="STARTING",
    )
    append_action(
        runs_dir=runs_dir,
        run_id=run_id,
        action="run_created",
        payload={"upload_id": upload_id, "input_path": input_path, "output_path": output_path},
    )
    return record


def write_run_record(*, runs_dir: Path, run_id: str, record: dict) -> dict:
    record["updated_at"] = int(time.time())
    path = run_record_path(runs_dir, run_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
    return record


def read_run_record(*, runs_dir: Path, run_id: str) -> dict | None:
    data = safe_read_json(run_record_path(runs_dir, run_id))
    return data if isinstance(data, dict) else None


def read_signoff(*, runs_dir: Path, run_id: str) -> dict | None:
    data = safe_read_json(signoff_path(runs_dir, run_id))
    return data if isinstance(data, dict) else None


def write_signoff(
    *,
    runs_dir: Path,
    run_id: str,
    decision: str,
    note: str,
    actor: str,
) -> dict:
    payload = {
        "run_id": run_id,
        "decision": decision,
        "note": note,
        "actor": actor,
        "timestamp": int(time.time()),
    }
    path = signoff_path(runs_dir, run_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    append_action(runs_dir=runs_dir, run_id=run_id, actor=actor, action="run_signoff", payload=payload)
    return payload


def read_review_state(*, runs_dir: Path, run_id: str) -> dict:
    data = safe_read_json(review_state_path(runs_dir, run_id))
    if isinstance(data, dict):
        return data
    return {"run_id": run_id, "rows": {}}


def write_review_state(*, runs_dir: Path, run_id: str, state: dict) -> dict:
    path = review_state_path(runs_dir, run_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    return state


def _project_review_state_from_run_rows(*, runs_dir: Path, run_id: str) -> dict:
    rows = fetch_run_rows(runs_dir=runs_dir, run_id=run_id, review_required_only=True)
    projected_rows: dict[str, dict] = {}
    for row in rows:
        row_index = int(row.get("row_index") or 0)
        if row_index <= 0:
            continue
        projected_rows[str(row_index)] = {
            "row_index": row_index,
            "item_number": str(row.get("item_number") or ""),
            "post_text": str(row.get("post_text") or ""),
            "assigned_category": str(row.get("assigned_category") or ""),
            "confidence_score": row.get("confidence_score"),
            "explanation": str(row.get("explanation") or ""),
            "flags": list(row.get("flags") or []),
            "fallback_events": list(row.get("fallback_events") or []),
            "soft_signal_score": row.get("soft_signal_score"),
            "soft_signal_flags": list(row.get("soft_signal_flags") or []),
            "soft_signal_evidence": list(row.get("soft_signal_evidence") or []),
            "review_required": True,
            "review_state": str(row.get("review_state") or "pending"),
            "review_decision": row.get("review_decision"),
            "reviewer_note": str(row.get("reviewer_note") or ""),
            "updated_at": int(row.get("updated_at") or time.time()),
        }
    return {"run_id": run_id, "rows": projected_rows}


def _review_state_from_canonical(
    *,
    runs_dir: Path,
    run_id: str,
    allow_checkpoint_backfill: bool = True,
) -> dict:
    state = _project_review_state_from_run_rows(runs_dir=runs_dir, run_id=run_id)
    if state.get("rows") or not allow_checkpoint_backfill:
        return state
    migrate_run_rows_to_canonical(runs_dir=runs_dir, run_id=run_id, include_checkpoints=True, include_output=False)
    return _project_review_state_from_run_rows(runs_dir=runs_dir, run_id=run_id)


def _checkpoint_rows_to_canonical(*, checkpoint_path: Path) -> list[dict]:
    rows: list[dict] = []
    for line in checkpoint_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            payload = json.loads(line)
        except Exception:
            continue
        if not isinstance(payload, dict):
            continue
        result = payload.get("result")
        if not isinstance(result, dict):
            continue
        flags = [str(item) for item in result.get("flags", []) if str(item)]
        try:
            row_index = int(result.get("row_index") or 0)
        except Exception:
            continue
        if row_index <= 0:
            continue
        model_votes = result.get("model_votes")
        rows.append(
            {
                "row_index": row_index,
                "row_hash": str(payload.get("row_hash") or ""),
                "assigned_category": str(result.get("category") or ""),
                "confidence_score": result.get("confidence"),
                "explanation": str(result.get("explanation") or ""),
                "flags": flags,
                "fallback_events": [str(item) for item in result.get("fallback_events", []) if str(item)],
                "soft_signal_score": result.get("soft_signal_score"),
                "soft_signal_flags": [str(item) for item in result.get("soft_signal_flags", []) if str(item)],
                "soft_signal_evidence": [str(item).strip() for item in result.get("soft_signal_evidence", []) if str(item).strip()],
                "judge_score": result.get("judge_score"),
                "judge_verdict": result.get("judge_verdict"),
                "consensus_tier": result.get("consensus_tier"),
                "minority_label": result.get("minority_label"),
                "model_votes": model_votes if isinstance(model_votes, dict) else None,
                "drafted_text": result.get("drafted_text"),
                "review_required": "REVIEW_REQUIRED" in flags,
            }
        )
    return rows


def _migrate_run_rows_from_checkpoints(*, runs_dir: Path, run_id: str) -> dict[str, int]:
    run_path = run_dir(runs_dir, run_id)
    record = read_run_record(runs_dir=runs_dir, run_id=run_id) or {}
    upload_id = str(record.get("upload_id")) if record.get("upload_id") else None
    attempt = latest_run_attempt(runs_dir=runs_dir, run_id=run_id)
    attempt_id = str(attempt.get("attempt_id") or "") if isinstance(attempt, dict) else ""
    checkpoint_paths = sorted(run_path.glob("_segments/*/runs/*/result_checkpoint.jsonl"))
    if not checkpoint_paths:
        return {"checkpoints_seen": 0, "rows_migrated": 0}

    checkpoint_rows: list[dict] = []
    for checkpoint_path in checkpoint_paths:
        checkpoint_rows.extend(_checkpoint_rows_to_canonical(checkpoint_path=checkpoint_path))

    if not checkpoint_rows:
        return {"checkpoints_seen": len(checkpoint_paths), "rows_migrated": 0}

    migrated = upsert_run_rows(
        runs_dir=runs_dir,
        run_id=run_id,
        upload_id=upload_id,
        attempt_id=attempt_id or None,
        rows=checkpoint_rows,
    )
    return {"checkpoints_seen": len(checkpoint_paths), "rows_migrated": migrated}


def _output_rows_to_canonical(*, output_path: Path) -> list[dict]:
    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_idx = {name: idx for idx, name in enumerate(header)}
        required = [
            "Item number",
            "Post text",
            "Assigned Category",
            "Confidence Score",
            "Explanation / Reasoning",
            "Flags",
            "Fallback Events",
            "Review Required",
        ]
        if any(name not in header_idx for name in required):
            return []

        rows: list[dict] = []
        for sheet_row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            original_row_index = row[header_idx[ORIGINAL_ROW_INDEX_COLUMN]] if ORIGINAL_ROW_INDEX_COLUMN in header_idx else None
            try:
                row_index = int(original_row_index or sheet_row_index)
            except Exception:
                row_index = sheet_row_index
            flags = "" if row[header_idx["Flags"]] is None else str(row[header_idx["Flags"]]).strip()
            fallback_events = "" if row[header_idx["Fallback Events"]] is None else str(row[header_idx["Fallback Events"]]).strip()
            soft_signal_flags = (
                "" if "Soft Signal Flags" not in header_idx or row[header_idx["Soft Signal Flags"]] is None else str(row[header_idx["Soft Signal Flags"]]).strip()
            )
            soft_signal_evidence = (
                ""
                if "Soft Signal Evidence" not in header_idx or row[header_idx["Soft Signal Evidence"]] is None
                else str(row[header_idx["Soft Signal Evidence"]]).strip()
            )
            model_votes = (
                ""
                if "Model Votes" not in header_idx or row[header_idx["Model Votes"]] is None
                else str(row[header_idx["Model Votes"]]).strip()
            )
            review_required = "" if row[header_idx["Review Required"]] is None else str(row[header_idx["Review Required"]]).strip()
            rows.append(
                {
                    "row_index": row_index,
                    "item_number": "" if row[header_idx["Item number"]] is None else str(row[header_idx["Item number"]]).strip(),
                    "post_text": "" if row[header_idx["Post text"]] is None else str(row[header_idx["Post text"]]).strip(),
                    "row_hash": "" if "Row Hash" not in header_idx or row[header_idx["Row Hash"]] is None else str(row[header_idx["Row Hash"]]).strip(),
                    "assigned_category": "" if row[header_idx["Assigned Category"]] is None else str(row[header_idx["Assigned Category"]]).strip(),
                    "confidence_score": row[header_idx["Confidence Score"]],
                    "explanation": "" if row[header_idx["Explanation / Reasoning"]] is None else str(row[header_idx["Explanation / Reasoning"]]).strip(),
                    "flags": [part for part in flags.split(";") if part],
                    "fallback_events": [part for part in fallback_events.split(";") if part],
                    "soft_signal_score": row[header_idx["Soft Signal Score"]] if "Soft Signal Score" in header_idx else None,
                    "soft_signal_flags": [part for part in soft_signal_flags.split(";") if part],
                    "soft_signal_evidence": [part.strip() for part in soft_signal_evidence.split("|") if part.strip()],
                    "judge_score": row[header_idx["Judge Score"]] if "Judge Score" in header_idx else None,
                    "judge_verdict": "" if "Judge Verdict" not in header_idx or row[header_idx["Judge Verdict"]] is None else str(row[header_idx["Judge Verdict"]]).strip(),
                    "consensus_tier": "" if "Consensus Tier" not in header_idx or row[header_idx["Consensus Tier"]] is None else str(row[header_idx["Consensus Tier"]]).strip(),
                    "minority_label": "" if "Minority Report" not in header_idx or row[header_idx["Minority Report"]] is None else str(row[header_idx["Minority Report"]]).strip(),
                    "model_votes": {"votes": [part.strip() for part in model_votes.split("|") if part.strip()]} if model_votes else None,
                    "review_required": review_required == "YES",
                }
            )
        return rows
    finally:
        wb.close()


def _migrate_run_rows_from_output(*, runs_dir: Path, run_id: str, output_path: Path) -> int:
    record = read_run_record(runs_dir=runs_dir, run_id=run_id) or {}
    upload_id = str(record.get("upload_id")) if record.get("upload_id") else None
    attempt = latest_run_attempt(runs_dir=runs_dir, run_id=run_id)
    attempt_id = str(attempt.get("attempt_id") or "") if isinstance(attempt, dict) else ""
    rows = _output_rows_to_canonical(output_path=output_path)
    if not rows:
        return 0
    return upsert_run_rows(
        runs_dir=runs_dir,
        run_id=run_id,
        upload_id=upload_id,
        attempt_id=attempt_id or None,
        rows=rows,
    )


def migrate_run_rows_to_canonical(
    *,
    runs_dir: Path,
    run_id: str,
    include_checkpoints: bool = True,
    include_output: bool = True,
) -> dict[str, int | bool | str]:
    record = read_run_record(runs_dir=runs_dir, run_id=run_id)
    if not record:
        return {}

    checkpoint_summary = {"checkpoints_seen": 0, "rows_migrated": 0}
    output_rows_migrated = 0
    output_present = False
    output_path = run_dir(runs_dir, run_id) / "output.xlsx"

    if include_checkpoints:
        checkpoint_summary = _migrate_run_rows_from_checkpoints(runs_dir=runs_dir, run_id=run_id)
    if include_output and output_path.exists():
        output_present = True
        output_rows_migrated = _migrate_run_rows_from_output(runs_dir=runs_dir, run_id=run_id, output_path=output_path)

    state = _project_review_state_from_run_rows(runs_dir=runs_dir, run_id=run_id)
    return {
        "run_id": run_id,
        "checkpoint_files_seen": int(checkpoint_summary["checkpoints_seen"]),
        "checkpoint_rows_migrated": int(checkpoint_summary["rows_migrated"]),
        "output_present": output_present,
        "output_rows_migrated": int(output_rows_migrated),
        "review_rows_projected": len((state.get("rows") or {})),
    }


def sync_review_rows_from_output(*, runs_dir: Path, run_id: str) -> dict:
    # Normal read paths now project review state from canonical `run_rows` and
    # only use checkpoint import as a temporary bridge for in-flight older runs.
    # Output workbook import remains an explicit migration command so reads stop
    # depending on parsing `output.xlsx` as live truth.
    state = _review_state_from_canonical(runs_dir=runs_dir, run_id=run_id, allow_checkpoint_backfill=True)
    return write_review_state(runs_dir=runs_dir, run_id=run_id, state=state)


def _pid_alive(pid: int | None) -> bool:
    from backend.services.run_lifecycle_service import pid_alive as lifecycle_pid_alive

    return lifecycle_pid_alive(pid)


def _pid_command(pid: int | None) -> str:
    from backend.services.run_lifecycle_service import pid_command as lifecycle_pid_command

    return lifecycle_pid_command(pid)


def _discover_run_process_pid(run_id: str) -> int | None:
    return discover_run_process_pid(run_id)


def _run_process_alive(run_id: str, pid: int | None) -> bool:
    if not _pid_alive(pid):
        return False
    command = _pid_command(pid)
    if not command or run_id not in command:
        return False
    return "backend/segment_worker.py" in command or "src.cli classify" in command


def _resolve_run_state(*, run_id: str, existing_state: str | None, progress: dict | list | None, control: dict | list | None) -> str:
    progress_state = str(progress.get("state")) if isinstance(progress, dict) and progress.get("state") else ""
    control_dict = control if isinstance(control, dict) else {}
    running = _run_process_alive(run_id, control_dict.get("pid"))
    paused = bool(control_dict.get("paused"))
    cancelled = bool(control_dict.get("cancelled") or control_dict.get("stopped_at"))
    nonterminal = {"STARTING", "PENDING", "VALIDATING", "PROCESSING", "WRITING", "PAUSED"}
    existing = str(existing_state or "UNKNOWN").upper()

    if cancelled and not running:
        return "CANCELLED"
    if paused and running:
        return "PAUSED"
    if progress_state in nonterminal and not running:
        return "INTERRUPTED"
    if existing in nonterminal and not running:
        return "INTERRUPTED"
    if progress_state:
        return progress_state
    return existing


def upsert_review_row(
    *,
    runs_dir: Path,
    run_id: str,
    row_index: int,
    review_state_value: str | None,
    review_decision: str | None,
    reviewer_note: str | None,
    actor: str,
) -> dict:
    update_run_row_review(
        runs_dir=runs_dir,
        run_id=run_id,
        row_index=row_index,
        review_state=review_state_value,
        review_decision=review_decision,
        reviewer_note=reviewer_note,
    )
    state = _project_review_state_from_run_rows(runs_dir=runs_dir, run_id=run_id)
    write_review_state(runs_dir=runs_dir, run_id=run_id, state=state)
    existing = state.get("rows", {}).get(str(row_index), {"row_index": row_index, "review_required": True})
    append_action(
        runs_dir=runs_dir,
        run_id=run_id,
        actor=actor,
        action="review_row_updated",
        payload={"row_index": row_index, "review_state": review_state_value, "review_decision": review_decision},
    )
    return existing


def refresh_run_record(*, runs_dir: Path, run_id: str, sync_review_rows: bool = True) -> dict | None:
    record = read_run_record(runs_dir=runs_dir, run_id=run_id)
    if not record:
        return None
    raw_progress = safe_read_json(run_dir(runs_dir, run_id) / "progress.json")
    processing_stats = safe_read_json(run_dir(runs_dir, run_id) / "processing_stats.json")
    control = safe_read_json(run_dir(runs_dir, run_id) / "control.json")
    control = control if isinstance(control, dict) else {}
    discovered_pid = _discover_run_process_pid(run_id)
    if discovered_pid and int(control.get("pid") or 0) != discovered_pid:
        control["pid"] = discovered_pid
        control.pop("shutdown_requested", None)
        control.pop("shutdown_requested_at", None)
        control.pop("shutdown_mode", None)
    elif _run_process_alive(run_id, control.get("pid")):
        control.pop("shutdown_requested", None)
        control.pop("shutdown_requested_at", None)
        control.pop("shutdown_mode", None)
    signoff = read_signoff(runs_dir=runs_dir, run_id=run_id)
    review_state = _project_review_state_from_run_rows(runs_dir=runs_dir, run_id=run_id)
    if sync_review_rows:
        write_review_state(runs_dir=runs_dir, run_id=run_id, state=review_state)

    rows = review_state.get("rows", {})
    detected_review = int(processing_stats.get("review_required_rows_detected") or 0) if isinstance(processing_stats, dict) else 0
    total_review = max(len(rows), detected_review)
    reviewed = sum(1 for row in rows.values() if row.get("review_state") not in {None, "", "pending"})
    record["state"] = _resolve_run_state(run_id=run_id, existing_state=record.get("state"), progress=raw_progress, control=control)
    segment_summary = build_run_segment_summary(runs_dir=runs_dir, run_id=run_id, effective_run_state=record["state"])
    record["state"] = normalize_state_from_segments(state=record["state"], segment_summary=segment_summary)
    canonical_stats = summarize_run_rows(runs_dir=runs_dir, run_id=run_id)
    total_rows = _resolved_total_rows(record=record, progress=raw_progress if isinstance(raw_progress, dict) else None, segment_summary=segment_summary)
    processed_rows = _resolved_processed_rows(
        record=record,
        progress=raw_progress if isinstance(raw_progress, dict) else None,
        canonical_stats=canonical_stats,
        segment_summary=segment_summary,
    )
    progress = dict(raw_progress) if isinstance(raw_progress, dict) else {}
    if str(progress.get("state") or "").upper() == "COMPLETED" and record["state"] != "COMPLETED":
        progress["state"] = record["state"]
    elif "state" not in progress:
        progress["state"] = record["state"]
    progress["processed_rows"] = processed_rows
    progress["total_rows"] = total_rows
    progress["progress_percentage"] = round((processed_rows / total_rows) * 100, 2) if total_rows > 0 else (100.0 if record["state"] == "COMPLETED" else 0.0)
    record["progress"] = progress
    base_processing_stats = processing_stats if isinstance(processing_stats, dict) else _fallback_processing_stats(record, progress)
    record["processing_stats"] = {
        **(base_processing_stats or {}),
        **canonical_stats,
        "processed_rows": processed_rows,
        "total_rows": total_rows,
    }
    record["processed_rows"] = processed_rows
    record["total_rows"] = total_rows
    record["progress_percentage"] = progress["progress_percentage"]
    record["row_progress_percentage"] = progress["progress_percentage"]
    record["control"] = control
    record["signoff"] = signoff
    record["review_summary"] = {
        "review_required_rows": total_review,
        "reviewed_rows": reviewed,
        "pending_rows": max(total_review - reviewed, 0),
    }
    persisted = write_run_record(runs_dir=runs_dir, run_id=run_id, record=record)
    update_run_snapshot(
        runs_dir=runs_dir,
        run_id=run_id,
        upload_id=str(record.get("upload_id")) if record.get("upload_id") else None,
        language=str(record.get("language")) if record.get("language") else None,
        review_mode=str(record.get("review_mode")) if record.get("review_mode") else None,
        state=str(record.get("state", "UNKNOWN")),
        progress=progress,
        created_at=int(record.get("created_at")) if record.get("created_at") else None,
    )
    update_latest_run_attempt_status(
        runs_dir=runs_dir,
        run_id=run_id,
        status=str(record.get("state", "UNKNOWN")),
    )
    return persisted


def list_run_records(*, runs_dir: Path) -> list[dict]:
    items = []
    if not runs_dir.exists():
        return items
    for child in sorted(runs_dir.iterdir(), reverse=True):
        if not child.is_dir() or child.name in {"uploads", RUN_HISTORY_DIRNAME}:
            continue
        record = refresh_run_record(runs_dir=runs_dir, run_id=child.name, sync_review_rows=False) or read_run_record(runs_dir=runs_dir, run_id=child.name)
        if record:
            items.append(record)
    return items


def _last_nonempty_log_line(path: Path) -> str | None:
    if not path.exists():
        return None
    try:
        for line in reversed(path.read_text(encoding="utf-8").splitlines()):
            stripped = line.strip()
            if stripped:
                return stripped
    except Exception:
        return None
    return None


def _run_log_excerpt(*, runs_dir: Path, run_id: str) -> str | None:
    candidates = sorted(
        [
            run_dir(runs_dir, run_id) / "logs.txt",
            runs_dir / f"{run_id}-segment-worker.log",
            runs_dir / f"{run_id}-classify-ui.log",
        ],
        key=lambda item: item.stat().st_mtime if item.exists() else 0,
        reverse=True,
    )
    for candidate in candidates:
        line = _last_nonempty_log_line(candidate)
        if line:
            return line
    return None


def _safe_path(value: object) -> Path | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return Path(str(value))
    except Exception:
        return None


def _path_is_within(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False


def _run_recovery_candidate(*, runs_dir: Path, run_id: str) -> dict | None:
    record = refresh_run_record(runs_dir=runs_dir, run_id=run_id, sync_review_rows=False)
    if not record:
        return None
    state = str(record.get("state") or "UNKNOWN").upper()
    if state in {"COMPLETED", "CANCELLED"}:
        return None

    control = record.get("control") if isinstance(record.get("control"), dict) else {}
    progress = record.get("progress") if isinstance(record.get("progress"), dict) else {}
    pid = control.get("pid")
    running = _run_process_alive(run_id, pid)
    paused = bool(control.get("paused"))
    upload_id = str(record.get("upload_id") or "") or None
    segment_summary = build_run_segment_summary(runs_dir=runs_dir, run_id=run_id, effective_run_state=state)
    pending_segments = int(segment_summary.get("segments_by_status", {}).get("QUEUED", 0))
    processing_segments = int(segment_summary.get("segments_by_status", {}).get("PROCESSING", 0))
    failed_segments = int(segment_summary.get("segments_by_status", {}).get("FAILED", 0))

    source_path: Path | None = None
    source_exists = False
    source_managed = False
    upload_record: dict | None = None
    broken_reasons: list[str] = []
    start_payload = record.get("start_payload") if isinstance(record.get("start_payload"), dict) else {}

    if upload_id:
        upload_record = read_upload_record(runs_dir=runs_dir, upload_id=upload_id)
        if not upload_record:
            broken_reasons.append("managed upload record is missing")
        else:
            source_path = _safe_path(upload_record.get("stored_path")) or _safe_path(record.get("input_path"))
            source_exists = bool(source_path and source_path.exists())
            source_managed = True
            if str(upload_record.get("status") or "").lower() != "accepted":
                broken_reasons.append("managed upload is no longer accepted")
            if source_path is None:
                broken_reasons.append("managed source workbook path is missing")
            elif not source_exists:
                broken_reasons.append("managed source workbook is missing from local storage")
    else:
        source_path = _safe_path(record.get("input_path"))
        source_exists = bool(source_path and source_path.exists())
        if source_path is None:
            broken_reasons.append("input workbook path is missing")
        elif not source_exists:
            broken_reasons.append("input workbook is missing")

    if state in {"FAILED", "INTERRUPTED", "STARTING", "PENDING", "VALIDATING", "PROCESSING", "WRITING", "PAUSED", "QUEUED"}:
        if not start_payload:
            broken_reasons.append("start payload is unavailable for restart")
        if not isinstance(progress, dict) or not progress:
            broken_reasons.append("progress record is missing")
        if upload_id and pending_segments <= 0 and processing_segments <= 0 and not running and state != "FAILED":
            broken_reasons.append("no resumable queued segments remain for this unfinished upload run")

    if running and state in {"FAILED", "INTERRUPTED"}:
        broken_reasons.append("run state is stale because a worker is still attached")

    continue_action: str | None = None
    if running and paused:
        continue_action = "resume"
    elif not running and source_exists and start_payload:
        if upload_id and (pending_segments > 0 or processing_segments > 0) and state in {"FAILED", "INTERRUPTED", "QUEUED", "PAUSED"}:
            continue_action = "recover"
        elif state in {"FAILED", "INTERRUPTED"}:
            continue_action = "retry"

    is_broken = bool(broken_reasons) and continue_action is None
    return {
        "run_id": run_id,
        "state": state,
        "upload_id": upload_id,
        "language": record.get("language"),
        "review_mode": record.get("review_mode"),
        "created_at": record.get("created_at"),
        "updated_at": record.get("updated_at"),
        "running": running,
        "paused": paused,
        "pid": pid if running else None,
        "progress": {
            "processed_rows": progress.get("processed_rows"),
            "total_rows": progress.get("total_rows"),
            "progress_percentage": progress.get("progress_percentage"),
        },
        "source_path": str(source_path) if source_path else None,
        "source_exists": source_exists,
        "source_managed": source_managed,
        "start_payload_available": bool(start_payload),
        "pending_segments": pending_segments,
        "processing_segments": processing_segments,
        "failed_segments": failed_segments,
        "can_continue": continue_action is not None,
        "continue_action": continue_action,
        "can_delete": not running,
        "is_broken": is_broken,
        "broken_reasons": broken_reasons,
        "last_log_line": _run_log_excerpt(runs_dir=runs_dir, run_id=run_id),
    }


def list_recovery_candidates(*, runs_dir: Path) -> list[dict]:
    candidates: list[dict] = []
    for record in list_run_records(runs_dir=runs_dir):
        run_id = str(record.get("run_id") or "")
        if not run_id:
            continue
        candidate = _run_recovery_candidate(runs_dir=runs_dir, run_id=run_id)
        if candidate:
            candidates.append(candidate)
    return sorted(
        candidates,
        key=lambda item: (
            0 if item.get("is_broken") else 1,
            0 if item.get("can_continue") else 1,
            -(int(item.get("updated_at") or 0)),
            str(item.get("run_id") or ""),
        ),
    )


def purge_run_state(*, runs_dir: Path, run_id: str, purge_source: bool = False) -> dict | None:
    record = refresh_run_record(runs_dir=runs_dir, run_id=run_id, sync_review_rows=False)
    if not record:
        return None
    control = record.get("control") if isinstance(record.get("control"), dict) else {}
    if _run_process_alive(run_id, control.get("pid")):
        raise RuntimeError("run is still active")

    run_path = run_dir(runs_dir, run_id)
    removed_paths: list[str] = []
    skipped_paths: list[str] = []
    upload_id = str(record.get("upload_id") or "") or None
    should_purge_upload = bool(
        purge_source
        and upload_id
        and count_other_runs_for_upload(runs_dir=runs_dir, upload_id=upload_id, excluding_run_id=run_id) == 0
    )

    if purge_source and upload_id and not should_purge_upload:
        skipped_paths.append(f"upload:{upload_id}")
    elif purge_source:
        input_path = _safe_path(record.get("input_path"))
        if input_path and input_path.exists() and _path_is_within(input_path, run_path):
            input_path.unlink(missing_ok=True)
            removed_paths.append(str(input_path))
        elif input_path:
            skipped_paths.append(str(input_path))

    if run_path.exists():
        import shutil

        shutil.rmtree(run_path)
        removed_paths.append(str(run_path))
    delete_run_snapshot(runs_dir=runs_dir, run_id=run_id)
    if should_purge_upload and upload_id:
        upload_path = runs_dir / UPLOADS_DIRNAME / upload_id
        if upload_path.exists():
            import shutil

            shutil.rmtree(upload_path)
            removed_paths.append(str(upload_path))
        delete_upload_snapshot(runs_dir=runs_dir, upload_id=upload_id)
    return {
        "run_id": run_id,
        "removed_paths": removed_paths,
        "skipped_paths": skipped_paths,
        "purged_source": bool(purge_source),
        "upload_id": upload_id,
    }


def build_run_detail(*, runs_dir: Path, run_id: str) -> dict | None:
    record = refresh_run_record(runs_dir=runs_dir, run_id=run_id, sync_review_rows=False)
    if not record:
        return None
    run_path = run_dir(runs_dir, run_id)
    progress = record.get("progress") or {}
    processing_stats = record.get("processing_stats") or {}
    signoff = record.get("signoff")
    review_state = _review_state_from_canonical(runs_dir=runs_dir, run_id=run_id, allow_checkpoint_backfill=False)
    review_rows = sorted(review_state.get("rows", {}).values(), key=lambda row: row.get("row_index", 0))
    review_required_target = max(
        len(review_rows),
        int(processing_stats.get("review_required_rows_detected") or 0) if isinstance(processing_stats, dict) else 0,
    )
    output_path = Path(str(record.get("output_path", ""))) if record.get("output_path") else None
    run_output_path = run_path / "output.xlsx"
    resolved_output = run_output_path if run_output_path.exists() else output_path

    artifacts = list_run_artifacts(runs_dir=runs_dir, run_id=run_id)
    control = record.get("control") or {}
    pid = control.get("pid")
    paused = bool(control.get("paused"))
    running = False
    if pid:
        running = _run_process_alive(run_id, pid)

    state = str(record.get("state", "UNKNOWN"))
    segment_summary = build_run_segment_summary(runs_dir=runs_dir, run_id=run_id, effective_run_state=state)
    pending_segments = int(segment_summary.get("segments_by_status", {}).get("QUEUED", 0))
    processing_segments = int(segment_summary.get("segments_by_status", {}).get("PROCESSING", 0))
    failed_segments = int(segment_summary.get("segments_by_status", {}).get("FAILED", 0))
    next_actions = []
    if state in {"FAILED", "INTERRUPTED"}:
        next_actions.extend(["inspect_failure", "retry_when_supported"])
    if state in {"COMPLETED"}:
        next_actions.append("download_output")
        if review_required_target > 0:
            next_actions.append("review_flagged_rows")
        else:
            next_actions.append("sign_off")
    if state in {"STARTING", "PENDING", "VALIDATING", "PROCESSING", "WRITING"}:
        next_actions.extend(["monitor_progress", "pause_or_stop_if_needed"])
    if state in {"INTERRUPTED", "FAILED"} and pending_segments > 0:
        next_actions.append("recover_segment_worker")
    if signoff:
        next_actions = ["view_signoff", "download_output"] + [x for x in next_actions if x != "sign_off"]
    next_actions = list(dict.fromkeys(next_actions))

    available_operations = {
        "pause": running and not paused and state in {"STARTING", "PENDING", "VALIDATING", "PROCESSING", "WRITING"},
        "resume": running and paused,
        "cancel": running,
        "retry": (not running) and state in {"FAILED", "CANCELLED", "INTERRUPTED"},
        "recover": True,
    }
    recovery = {
        "running": running,
        "paused": paused,
        "pid": pid,
        "has_control": bool(control),
        "has_progress": isinstance(progress, dict),
        "output_ready": bool(resolved_output and Path(resolved_output).exists()),
        "can_retry": available_operations["retry"],
        "can_cancel": available_operations["cancel"],
        "can_resume_worker": (not running) and bool(record.get("upload_id")) and pending_segments > 0,
        "pending_segments": pending_segments,
        "processing_segments": processing_segments,
        "failed_segments": failed_segments,
    }

    return {
        "run_id": run_id,
        "state": state,
        "language": record.get("language"),
        "review_mode": record.get("review_mode"),
        "upload_id": record.get("upload_id"),
        "input_path": record.get("input_path"),
        "output_path": str(resolved_output) if resolved_output else record.get("output_path"),
        "output_ready": bool(resolved_output and Path(resolved_output).exists()),
        "created_at": record.get("created_at"),
        "updated_at": record.get("updated_at"),
        "progress": {
            "state": progress.get("state"),
            "message": progress.get("message"),
            "total_rows": progress.get("total_rows"),
            "processed_rows": progress.get("processed_rows"),
            "progress_percentage": progress.get("progress_percentage"),
            "started_at": progress.get("started_at"),
            "completed_at": progress.get("completed_at"),
        },
        "processing_stats": processing_stats,
        "review_summary": record.get("review_summary"),
        "signoff": signoff,
        "next_actions": next_actions,
        "available_operations": available_operations,
        "recovery": recovery,
        "segment_summary": segment_summary,
        "review_rows_preview": review_rows[:10],
        "artifacts": artifacts,
    }


def list_run_artifacts(*, runs_dir: Path, run_id: str) -> list[dict]:
    run_path = run_dir(runs_dir, run_id)
    artifacts = []
    for name in ARTIFACT_NAMES:
        p = run_path / name
        if p.exists():
            artifacts.append({"name": name, "path": str(p), "bytes": p.stat().st_size})
    return artifacts


def build_artifact_center(*, runs_dir: Path, run_id: str) -> dict | None:
    record = refresh_run_record(runs_dir=runs_dir, run_id=run_id, sync_review_rows=False)
    if not record:
        return None
    artifacts = list_run_artifacts(runs_dir=runs_dir, run_id=run_id)
    items = []
    for item in artifacts:
        purpose = {
            "output.xlsx": "Governed output workbook",
            "integrity_report.json": "Integrity and distribution audit report",
            "artifact_manifest.json": "Artifact hash manifest",
            "segment_assignment_manifest.jsonl": "Row-to-segment assignment ledger for audit and recovery",
            "policy.json": "Resolved run policy and route metadata",
            "logs.txt": "Execution log for the run lifecycle",
            "progress.json": "Lifecycle progress record",
            "processing_stats.json": "Live row/time/threat snapshot during processing",
            "review_state.json": "Persistent browser review state",
            "signoff.json": "Acceptance decision record",
            "action_log.jsonl": "Operator action log",
            "disagreement_report.json": "Disagreement evidence when produced",
            "control.json": "Local process control state",
        }.get(item["name"], "Run artifact")
        items.append(
            {
                **item,
                "purpose": purpose,
                "download_path": f"/runs/{run_id}/artifacts/download/{item['name']}",
            }
        )
    return {
        "run_id": run_id,
        "state": record.get("state"),
        "signoff": record.get("signoff"),
        "review_summary": record.get("review_summary"),
        "artifacts": items,
    }


def build_review_queue(
    *,
    runs_dir: Path,
    run_id: str,
    review_state_filter: str | None = None,
    decision_filter: str | None = None,
    sort_by: str = "row_index",
    sort_order: str = "asc",
) -> dict | None:
    record = refresh_run_record(runs_dir=runs_dir, run_id=run_id, sync_review_rows=False)
    if not record:
        return None
    state = _review_state_from_canonical(runs_dir=runs_dir, run_id=run_id, allow_checkpoint_backfill=False)
    rows = list(state.get("rows", {}).values())

    if review_state_filter and review_state_filter != "all":
        rows = [row for row in rows if str(row.get("review_state", "pending")) == review_state_filter]
    if decision_filter and decision_filter != "all":
        rows = [row for row in rows if str(row.get("review_decision") or "") == decision_filter]

    reverse = sort_order == "desc"
    if sort_by == "confidence":
        rows.sort(key=lambda row: float(row.get("confidence_score") or 0), reverse=reverse)
    elif sort_by == "category":
        rows.sort(key=lambda row: str(row.get("assigned_category") or ""), reverse=reverse)
    elif sort_by == "review_state":
        rows.sort(key=lambda row: str(row.get("review_state") or "pending"), reverse=reverse)
    else:
        rows.sort(key=lambda row: int(row.get("row_index") or 0), reverse=reverse)

    return {
        "run_id": run_id,
        "state": record.get("state"),
        "review_summary": record.get("review_summary"),
        "filters": {
            "review_state": review_state_filter or "all",
            "review_decision": decision_filter or "all",
            "sort_by": sort_by,
            "sort_order": sort_order,
        },
        "rows": rows,
    }


def build_row_inspector(*, runs_dir: Path, run_id: str, row_index: int) -> dict | None:
    record = refresh_run_record(runs_dir=runs_dir, run_id=run_id, sync_review_rows=False)
    if not record:
        return None
    state = _review_state_from_canonical(runs_dir=runs_dir, run_id=run_id, allow_checkpoint_backfill=False)
    row = fetch_run_row(runs_dir=runs_dir, run_id=run_id, row_index=row_index)
    if row:
        row = {
            "row_index": int(row.get("row_index") or row_index),
            "item_number": str(row.get("item_number") or ""),
            "post_text": str(row.get("post_text") or ""),
            "assigned_category": str(row.get("assigned_category") or ""),
            "confidence_score": row.get("confidence_score"),
            "explanation": str(row.get("explanation") or ""),
            "flags": list(row.get("flags") or []),
            "fallback_events": list(row.get("fallback_events") or []),
            "soft_signal_score": row.get("soft_signal_score"),
            "soft_signal_flags": list(row.get("soft_signal_flags") or []),
            "soft_signal_evidence": list(row.get("soft_signal_evidence") or []),
            "review_required": bool(row.get("review_required")),
            "review_state": str(row.get("review_state") or "pending"),
            "review_decision": row.get("review_decision"),
            "reviewer_note": str(row.get("reviewer_note") or ""),
            "updated_at": int(row.get("updated_at") or time.time()),
        }
    else:
        row = state.get("rows", {}).get(str(row_index))
    if not row:
        return None
    disagreement_path = run_dir(runs_dir, run_id) / "disagreement_report.json"
    disagreement_row = None
    if disagreement_path.exists():
        try:
            payload = json.loads(disagreement_path.read_text(encoding="utf-8"))
            for item in payload.get("rows", []):
                if int(item.get("row_index", -1)) == row_index:
                    disagreement_row = item
                    break
        except Exception:
            disagreement_row = None
    return {
        "run_id": run_id,
        "row_index": row_index,
        "run_state": record.get("state"),
        "language": record.get("language"),
        "review_mode": record.get("review_mode"),
        "row": row,
        "evidence": {
            "explanation": row.get("explanation"),
            "flags": row.get("flags", []),
            "fallback_events": row.get("fallback_events", []),
            "soft_signal_score": row.get("soft_signal_score"),
            "soft_signal_flags": row.get("soft_signal_flags", []),
            "soft_signal_evidence": row.get("soft_signal_evidence", []),
            "disagreement": disagreement_row,
        },
        "review_controls": {
            "review_state": row.get("review_state", "pending"),
            "review_decision": row.get("review_decision"),
            "reviewer_note": row.get("reviewer_note", ""),
        },
    }
