from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import backend.main as main_module
from backend.segment_worker import _reset_segment_attempt_artifacts, _should_resume_segment_attempt
from backend.segment_worker import _rebuild_segment_output_from_canonical
from backend.services import auth_service
from backend.services.ops_db_service import (
    build_operations_overview,
    build_run_segment_summary,
    build_upload_queue_summary,
    claim_next_segment,
    complete_segment,
    fetch_run_row,
    reconcile_run_segments,
    record_upload,
    register_run,
    resume_run_segments,
    upsert_run_rows,
    update_run_snapshot,
    update_segment_progress,
)
from backend.services.run_state_service import create_run_record
from backend.services.run_state_service import sync_review_rows_from_output
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from src.classifier import stable_row_hash


class OpsQueueRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.runs_dir = Path(self.tempdir.name) / "runs"
        self.runs_dir.mkdir(parents=True, exist_ok=True)
        self.previous_runs_dir = main_module.RUNS_DIR
        main_module.RUNS_DIR = self.runs_dir
        auth_service._SESSIONS.clear()
        self.client = TestClient(main_module.app)

    def tearDown(self) -> None:
        main_module.RUNS_DIR = self.previous_runs_dir
        auth_service._SESSIONS.clear()
        self.tempdir.cleanup()

    def _login_admin(self) -> None:
        response = self.client.post(
            "/auth/login",
            json={"actor_name": "ops-queue-test", "role": "admin", "access_code": "spot-local"},
        )
        self.assertEqual(response.status_code, 200, response.text)

    def test_operations_overview_tracks_segment_queue_and_progress(self) -> None:
        self._login_admin()
        upload_response = self.client.post(
            "/uploads/intake?filename=queue-check.xlsx",
            content=(
                b"PK\x03\x04"
                b"synthetic-invalid-xlsx"
            ),
        )
        self.assertEqual(upload_response.status_code, 200, upload_response.text)
        rejected = upload_response.json()
        self.assertEqual(rejected["status"], "rejected")

        accepted_upload_id = "upload-segmented"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "segmented.xlsx",
            "stored_path": str(upload_dir / "segmented.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000000,
            "validation": {"accepted": True, "row_count": 1200},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")

        from backend.services.excel_service import read_upload_record

        record_upload(runs_dir=self.runs_dir, record=record)
        queued_upload = read_upload_record(runs_dir=self.runs_dir, upload_id=accepted_upload_id)
        self.assertEqual((queued_upload or {}).get("queue_summary", {}).get("segment_count"), 3)
        self.assertEqual((queued_upload or {}).get("queue_summary", {}).get("segments_by_status", {}).get("READY"), 3)
        self.assertEqual((queued_upload or {}).get("queue_summary", {}).get("row_progress_percentage"), 0.0)
        self.assertEqual((queued_upload or {}).get("queue_summary", {}).get("segment_progress_percentage"), 0.0)

        run_id = "queue-run-001"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(upload_dir / "segmented.xlsx"),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": accepted_upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        first_segment = claim_next_segment(runs_dir=self.runs_dir, run_id=run_id, worker_id="queue-worker")
        self.assertIsNotNone(first_segment)
        update_segment_progress(runs_dir=self.runs_dir, segment_id=str(first_segment["segment_id"]), processed_rows=500)
        complete_segment(runs_dir=self.runs_dir, segment_id=str(first_segment["segment_id"]))
        second_segment = claim_next_segment(runs_dir=self.runs_dir, run_id=run_id, worker_id="queue-worker")
        self.assertIsNotNone(second_segment)
        update_segment_progress(runs_dir=self.runs_dir, segment_id=str(second_segment["segment_id"]), processed_rows=200)
        update_run_snapshot(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
            progress={"run_id": run_id, "state": "PROCESSING", "processed_rows": 700, "total_rows": 1200, "progress_percentage": 58.33},
            created_at=1770000000,
        )

        overview = build_operations_overview(runs_dir=self.runs_dir)
        self.assertEqual(overview["uploads"], 2)
        self.assertEqual(overview["total_segments"], 3)
        self.assertEqual(overview["total_rows"], 1200)
        self.assertEqual(overview["processed_rows"], 700)
        self.assertEqual(overview["progress_percentage"], 58.33)
        self.assertEqual(overview["segments_by_status"]["COMPLETED"], 1)
        self.assertEqual(overview["segments_by_status"]["PROCESSING"], 1)
        self.assertEqual(overview["segments_by_status"]["QUEUED"], 1)

        api_response = self.client.get("/operations/overview")
        self.assertEqual(api_response.status_code, 200, api_response.text)
        api_data = api_response.json()
        self.assertEqual(api_data["total_segments"], 3)
        segmented_summary = next(item for item in api_data["recent_uploads"] if item["upload_id"] == accepted_upload_id)
        self.assertEqual(segmented_summary["segment_count"], 3)
        self.assertEqual(segmented_summary["row_progress_percentage"], 58.33)
        self.assertEqual(segmented_summary["segment_progress_percentage"], 33.33)

    def test_historical_run_without_indexed_upload_does_not_break_runs_listing(self) -> None:
        self._login_admin()
        run_id = "historical-run-no-upload"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path="samples/sample_germany.xlsx",
            output_path=str(run_dir / "output.xlsx"),
            upload_id="missing-upload-id",
            language="de",
            review_mode="partial",
            start_payload={"upload_id": "missing-upload-id", "language": "de", "review_mode": "partial"},
        )
        (run_dir / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "COMPLETED", "processed_rows": 1, "total_rows": 1, "progress_percentage": 100.0}),
            encoding="utf-8",
        )
        response = self.client.get("/runs")
        self.assertEqual(response.status_code, 200, response.text)
        data = response.json()
        matched = next(item for item in data if item["run_id"] == run_id)
        self.assertEqual(matched["run_id"], run_id)

    def test_reconcile_run_segments_returns_orphaned_processing_work_to_queue(self) -> None:
        accepted_upload_id = "upload-reconcile"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "reconcile.xlsx",
            "stored_path": str(upload_dir / "reconcile.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000100,
            "validation": {"accepted": True, "row_count": 1200},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)
        register_run(
            runs_dir=self.runs_dir,
            run_id="reconcile-run",
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        first_segment = claim_next_segment(runs_dir=self.runs_dir, run_id="reconcile-run", worker_id="worker-a")
        self.assertIsNotNone(first_segment)
        update_segment_progress(runs_dir=self.runs_dir, segment_id=str(first_segment["segment_id"]), processed_rows=125)
        reconciled = reconcile_run_segments(runs_dir=self.runs_dir, run_id="reconcile-run", target_state="QUEUED")
        self.assertEqual(reconciled, 1)
        overview = build_operations_overview(runs_dir=self.runs_dir)
        self.assertEqual(overview["segments_by_status"]["QUEUED"], 3)
        self.assertEqual(overview["processed_rows"], 125)
        self.assertEqual(overview["progress_percentage"], 10.42)
        summary = next(item for item in overview["recent_uploads"] if item["upload_id"] == accepted_upload_id)
        self.assertEqual(summary["row_progress_percentage"], 10.42)
        self.assertEqual(summary["segment_progress_percentage"], 0.0)

    def test_segment_attempt_reset_clears_stale_child_artifacts(self) -> None:
        segment_dir = self.runs_dir / "run-x" / "_segments" / "seg-00001"
        segment_runs_dir = segment_dir / "runs"
        child_run_dir = segment_runs_dir / "run-x-segment-00001"
        child_run_dir.mkdir(parents=True, exist_ok=True)

        (segment_dir / "output.xlsx").write_text("stale-output", encoding="utf-8")
        (segment_dir / "worker.log").write_text("stale-log", encoding="utf-8")
        (child_run_dir / "progress.json").write_text(json.dumps({"processed_rows": 30}), encoding="utf-8")
        (child_run_dir / "processing_stats.json").write_text(json.dumps({"processed_rows": 30}), encoding="utf-8")

        _reset_segment_attempt_artifacts(
            segment_dir=segment_dir,
            segment_runs_dir=segment_runs_dir,
            segment_run_id="run-x-segment-00001",
        )

        self.assertFalse((segment_dir / "output.xlsx").exists())
        self.assertFalse((segment_dir / "worker.log").exists())
        self.assertFalse(child_run_dir.exists())

    def test_segment_attempt_reset_preserves_child_run_when_resuming(self) -> None:
        segment_dir = self.runs_dir / "run-y" / "_segments" / "seg-00001"
        segment_runs_dir = segment_dir / "runs"
        child_run_dir = segment_runs_dir / "run-y-segment-00001"
        child_run_dir.mkdir(parents=True, exist_ok=True)

        (segment_dir / "output.xlsx").write_text("stale-output", encoding="utf-8")
        (segment_dir / "worker.log").write_text("stale-log", encoding="utf-8")
        (child_run_dir / "progress.json").write_text(json.dumps({"processed_rows": 30}), encoding="utf-8")

        _reset_segment_attempt_artifacts(
            segment_dir=segment_dir,
            segment_runs_dir=segment_runs_dir,
            segment_run_id="run-y-segment-00001",
            preserve_child_run=True,
        )

        self.assertFalse((segment_dir / "output.xlsx").exists())
        self.assertFalse((segment_dir / "worker.log").exists())
        self.assertTrue(child_run_dir.exists())

    def test_resume_segment_attempt_uses_committed_progress_not_checkpoint_presence(self) -> None:
        self.assertFalse(_should_resume_segment_attempt(resume_existing=False, committed_processed_rows=10))
        self.assertFalse(_should_resume_segment_attempt(resume_existing=True, committed_processed_rows=0))
        self.assertTrue(_should_resume_segment_attempt(resume_existing=True, committed_processed_rows=1))

    def test_rebuild_segment_output_from_canonical_restores_missing_completed_segment_output(self) -> None:
        upload_id = "upload-rebuild-completed-segment"
        upload_dir = self.runs_dir / "uploads" / upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = upload_dir / "segment-source.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Item number", "Post text", "Category"])
        ws.append(["1", "First segment row", ""])
        ws.append(["2", "Second segment row", ""])
        wb.save(workbook_path)
        wb.close()

        record = {
            "upload_id": upload_id,
            "filename": "segment-source.xlsx",
            "stored_path": str(workbook_path),
            "bytes": workbook_path.stat().st_size,
            "status": "accepted",
            "created_at": 1770000500,
            "validation": {"accepted": True, "row_count": 2},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)

        from backend.services.ops_db_service import replace_upload_rows

        replace_upload_rows(
            runs_dir=self.runs_dir,
            upload_id=upload_id,
            rows=[
                {
                    "sequence_index": 1,
                    "row_index": 2,
                    "item_number": "1",
                    "post_text": "First segment row",
                    "source_category": "",
                    "row_hash": stable_row_hash("1", "First segment row"),
                    "post_text_sha256": "sha-segment-1",
                    "post_text_length": 17,
                },
                {
                    "sequence_index": 2,
                    "row_index": 3,
                    "item_number": "2",
                    "post_text": "Second segment row",
                    "source_category": "",
                    "row_hash": stable_row_hash("2", "Second segment row"),
                    "post_text_sha256": "sha-segment-2",
                    "post_text_length": 18,
                },
            ],
        )

        run_id = "rebuild-completed-segment-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(workbook_path),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        segment = claim_next_segment(runs_dir=self.runs_dir, run_id=run_id, worker_id="worker-segment-rebuild")
        self.assertIsNotNone(segment)
        complete_segment(runs_dir=self.runs_dir, segment_id=str(segment["segment_id"]))
        upsert_run_rows(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=upload_id,
            rows=[
                {
                    "row_index": 2,
                    "row_hash": stable_row_hash("1", "First segment row"),
                    "assigned_category": "Not Antisemitic",
                    "confidence_score": 0.91,
                    "explanation": "Recovered first row.",
                    "flags": [],
                    "review_required": False,
                },
                {
                    "row_index": 3,
                    "row_hash": stable_row_hash("2", "Second segment row"),
                    "assigned_category": "Anti-Israel",
                    "confidence_score": 0.87,
                    "explanation": "Recovered second row.",
                    "flags": ["REVIEW_REQUIRED"],
                    "review_required": True,
                },
            ],
        )

        segment_dir = run_dir / "_segments" / str(segment["segment_id"])
        rebuilt_rows = _rebuild_segment_output_from_canonical(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=upload_id,
            segment_id=str(segment["segment_id"]),
            segment_run_id=f"{run_id}-segment-00001",
            segment_input_path=segment_dir / "input.xlsx",
            segment_output_path=segment_dir / "output.xlsx",
            segment_dir=segment_dir,
            row_start=int(segment["row_start"]),
            row_end=int(segment["row_end"]),
            row_count=int(segment["row_count"]),
            language="de",
            review_mode="partial",
            ssot_path=PROJECT_ROOT / "ssot/ssot.json",
        )
        self.assertEqual(rebuilt_rows, 2)
        self.assertTrue((segment_dir / "input.xlsx").exists())
        self.assertTrue((segment_dir / "output.xlsx").exists())

        rebuilt_wb = load_workbook(segment_dir / "output.xlsx", read_only=True, data_only=True)
        try:
            rebuilt_ws = rebuilt_wb[rebuilt_wb.sheetnames[0]]
            header = [str(v).strip() if v is not None else "" for v in next(rebuilt_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            assigned_category_idx = header.index("Assigned Category")
            rows = list(rebuilt_ws.iter_rows(min_row=2, max_row=3, values_only=True))
        finally:
            rebuilt_wb.close()
        self.assertEqual(rows[0][assigned_category_idx], "Not Antisemitic")
        self.assertEqual(rows[1][assigned_category_idx], "Anti-Israel")

    def test_resume_run_segments_requeues_failed_work_without_dropping_progress(self) -> None:
        accepted_upload_id = "upload-resume"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "resume.xlsx",
            "stored_path": str(upload_dir / "resume.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000100,
            "validation": {"accepted": True, "row_count": 1200},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)
        register_run(
            runs_dir=self.runs_dir,
            run_id="resume-run",
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="FAILED",
        )
        first_segment = claim_next_segment(runs_dir=self.runs_dir, run_id="resume-run", worker_id="worker-a")
        self.assertIsNotNone(first_segment)
        update_segment_progress(runs_dir=self.runs_dir, segment_id=str(first_segment["segment_id"]), processed_rows=125)
        from backend.services.ops_db_service import fail_segment

        fail_segment(runs_dir=self.runs_dir, segment_id=str(first_segment["segment_id"]), error_message="synthetic")
        resumed = resume_run_segments(runs_dir=self.runs_dir, run_id="resume-run")
        self.assertEqual(resumed, 3)
        overview = build_operations_overview(runs_dir=self.runs_dir)
        self.assertEqual(overview["processed_rows"], 125)
        self.assertEqual(overview["segments_by_status"]["QUEUED"], 3)

    def test_operations_overview_does_not_count_stale_starting_run_as_active(self) -> None:
        accepted_upload_id = "upload-stale-starting"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "stale-starting.xlsx",
            "stored_path": str(upload_dir / "stale-starting.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000200,
            "validation": {"accepted": True, "row_count": 50},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)

        run_id = "stale-starting-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(upload_dir / "stale-starting.xlsx"),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": accepted_upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="STARTING",
        )

        overview = build_operations_overview(runs_dir=self.runs_dir)
        self.assertEqual(overview["active_uploads"], 0)
        summary = next(item for item in overview["recent_uploads"] if item["upload_id"] == accepted_upload_id)
        self.assertEqual((summary.get("run") or {}).get("state"), "INTERRUPTED")

    def test_sync_review_rows_from_checkpoints_surfaces_live_rows(self) -> None:
        accepted_upload_id = "upload-live-review"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "live-review.xlsx",
            "stored_path": str(upload_dir / "live-review.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000300,
            "validation": {"accepted": True, "row_count": 2},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)
        from backend.services.ops_db_service import replace_upload_rows

        replace_upload_rows(
            runs_dir=self.runs_dir,
            upload_id=accepted_upload_id,
            rows=[
                {
                    "sequence_index": 1,
                    "row_index": 2,
                    "item_number": "1",
                    "post_text": "Checkpoint review row",
                    "source_category": "",
                    "row_hash": "hash-1",
                    "post_text_sha256": "sha-1",
                    "post_text_length": 21,
                }
            ],
        )
        run_id = "live-review-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(upload_dir / "live-review.xlsx"),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": accepted_upload_id, "language": "de", "review_mode": "partial"},
        )
        checkpoint_dir = run_dir / "_segments" / f"{accepted_upload_id}-seg-00001" / "runs" / f"{run_id}-segment-00001"
        checkpoint_dir.mkdir(parents=True, exist_ok=True)
        (checkpoint_dir / "result_checkpoint.jsonl").write_text(
            json.dumps(
                {
                    "result_index": 0,
                    "row_hash": "hash-1",
                    "result": {
                        "row_index": 2,
                        "raw_category": "Anti-Israel",
                        "category": "Anti-Israel",
                        "confidence": 0.88,
                        "explanation": "Live checkpointed row.",
                        "flags": ["REVIEW_REQUIRED"],
                        "soft_signal_score": 0.0,
                        "soft_signal_flags": [],
                        "soft_signal_evidence": [],
                        "resolved_model_version": "synthetic-model",
                        "model_votes": {},
                        "consensus_tier": None,
                        "minority_label": None,
                        "drafted_text": None,
                        "judge_score": None,
                        "judge_verdict": None,
                        "fallback_events": [],
                    },
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8",
        )
        state = sync_review_rows_from_output(runs_dir=self.runs_dir, run_id=run_id)
        self.assertEqual(len(state.get("rows", {})), 1)
        row = state["rows"]["2"]
        self.assertEqual(row["assigned_category"], "Anti-Israel")
        self.assertEqual(row["post_text"], "Checkpoint review row")
        stored = fetch_run_row(runs_dir=self.runs_dir, run_id=run_id, row_index=2)
        self.assertIsNotNone(stored)
        self.assertEqual((stored or {}).get("assigned_category"), "Anti-Israel")
        self.assertTrue((stored or {}).get("review_required"))

    def test_upload_queue_summary_prefers_canonical_run_rows_for_live_progress(self) -> None:
        accepted_upload_id = "upload-canonical-progress"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "canonical-progress.xlsx",
            "stored_path": str(upload_dir / "canonical-progress.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000400,
            "validation": {"accepted": True, "row_count": 2},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)
        from backend.services.ops_db_service import replace_upload_rows

        replace_upload_rows(
            runs_dir=self.runs_dir,
            upload_id=accepted_upload_id,
            rows=[
                {
                    "sequence_index": 1,
                    "row_index": 2,
                    "item_number": "1",
                    "post_text": "Canonical progress row",
                    "source_category": "",
                    "row_hash": "hash-progress-1",
                    "post_text_sha256": "sha-progress-1",
                    "post_text_length": 22,
                },
                {
                    "sequence_index": 2,
                    "row_index": 3,
                    "item_number": "2",
                    "post_text": "Canonical progress row 2",
                    "source_category": "",
                    "row_hash": "hash-progress-2",
                    "post_text_sha256": "sha-progress-2",
                    "post_text_length": 24,
                },
            ],
        )
        run_id = "canonical-progress-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(upload_dir / "canonical-progress.xlsx"),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": accepted_upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        upsert_run_rows(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=accepted_upload_id,
            rows=[
                {
                    "row_index": 2,
                    "assigned_category": "Anti-Israel",
                    "confidence_score": 0.9,
                    "explanation": "Committed canonical row.",
                    "flags": ["REVIEW_REQUIRED"],
                    "review_required": True,
                }
            ],
        )

        summary = build_upload_queue_summary(runs_dir=self.runs_dir, upload_id=accepted_upload_id)
        self.assertEqual(summary["processed_rows"], 1)
        self.assertEqual(summary["processing_stats"]["processed_rows"], 1)
        self.assertEqual(summary["processing_stats"]["review_required_rows_detected"], 1)

    def test_run_detail_read_does_not_reconcile_segments_on_interrupt(self) -> None:
        self._login_admin()
        accepted_upload_id = "upload-read-no-reconcile"
        upload_dir = self.runs_dir / "uploads" / accepted_upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "upload_id": accepted_upload_id,
            "filename": "read-no-reconcile.xlsx",
            "stored_path": str(upload_dir / "read-no-reconcile.xlsx"),
            "bytes": 1024,
            "status": "accepted",
            "created_at": 1770000500,
            "validation": {"accepted": True, "row_count": 1200},
        }
        (upload_dir / "upload.json").write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=record)

        run_id = "interrupt-read-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(upload_dir / "read-no-reconcile.xlsx"),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": accepted_upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=accepted_upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        first_segment = claim_next_segment(runs_dir=self.runs_dir, run_id=run_id, worker_id="worker-a")
        self.assertIsNotNone(first_segment)
        (run_dir / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "PROCESSING", "processed_rows": 0, "total_rows": 1200, "progress_percentage": 0.0}),
            encoding="utf-8",
        )
        (run_dir / "control.json").write_text(json.dumps({"run_id": run_id, "pid": None, "paused": False}), encoding="utf-8")

        response = self.client.get(f"/runs/{run_id}/detail")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["state"], "INTERRUPTED")
        self.assertEqual(payload["segment_summary"]["segments_by_status"]["QUEUED"], 3)

        raw_summary = build_run_segment_summary(runs_dir=self.runs_dir, run_id=run_id)
        self.assertEqual(raw_summary["segments_by_status"]["PROCESSING"], 1)


if __name__ == "__main__":
    suite = unittest.defaultTestLoader.loadTestsFromTestCase(OpsQueueRegressionTests)
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    raise SystemExit(0 if result.wasSuccessful() else 1)
