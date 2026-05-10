from __future__ import annotations

import io
import json
import shutil
import sys
import tempfile
import unittest
import uuid
from unittest.mock import patch
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook

import backend.main as main_module
from backend.services import auth_service
from backend.services.excel_service import ensure_upload_rows_materialized, intake_workbook
from backend.services.artifact_manifest_service import build_artifact_manifest
from backend.services.ops_db_service import fetch_run_row, fetch_upload_rows_for_segment, start_run_attempt
from backend.services.run_state_service import build_review_queue, create_run_record
from backend.services import run_state_service
from src.excel_io import build_segment_input_workbook_from_entries, read_input_rows
from src.classifier import stable_row_hash
from src.models import ClassificationResult
from src.pipeline import _append_result_checkpoint, run_classification
from src.ssot_loader import load_ssot


def _build_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Item number", "Post text", "Category"])
    ws.append(["1", "Synthetic contract test row", ""])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _prepare_synthetic_run(runs_dir: Path, run_id: str) -> Path:
    run_path = runs_dir / run_id
    if run_path.exists():
        shutil.rmtree(run_path)
    run_path.mkdir(parents=True, exist_ok=True)

    create_run_record(
        runs_dir=runs_dir,
        run_id=run_id,
        input_path="samples/sample_germany.xlsx",
        output_path=str(run_path / "output.xlsx"),
        upload_id=None,
        language="de",
        review_mode="partial",
        start_payload={"upload_id": "synthetic-upload", "language": "de", "review_mode": "partial", "limit": 1},
    )

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Item number",
            "Post text",
            "Category",
            "Assigned Category",
            "Fallback Events",
            "Confidence Score",
            "Explanation / Reasoning",
            "Flags",
            "Review Required",
        ]
    )
    ws.append(
        [
            "1",
            "Synthetic flagged row for contract regression.",
            "",
            "Anti-Israel",
            "CLASSIFIER_FALLBACK_FAILED",
            0.41,
            "Synthetic reasoning.",
            "LOW_CONFIDENCE",
            "YES",
        ]
    )
    wb.save(run_path / "output.xlsx")
    wb.close()

    (run_path / "progress.json").write_text(
        json.dumps({"run_id": run_id, "state": "COMPLETED", "progress_percentage": 100.0, "processed_rows": 1, "total_rows": 1}),
        encoding="utf-8",
    )
    (run_path / "control.json").write_text(
        json.dumps({"run_id": run_id, "pid": None, "paused": False, "output": str(run_path / "output.xlsx")}),
        encoding="utf-8",
    )
    (run_path / "artifact_manifest.json").write_text(json.dumps({"artifacts": ["output.xlsx"]}), encoding="utf-8")
    (run_path / "integrity_report.json").write_text(json.dumps({"status": "ok"}), encoding="utf-8")
    (run_path / "policy.json").write_text(json.dumps({"policy": "synthetic"}), encoding="utf-8")
    (run_path / "logs.txt").write_text("synthetic contract log\n", encoding="utf-8")
    return run_path


class BackendContractRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.runs_dir = Path(self.tempdir.name) / "runs"
        self.runs_dir.mkdir(parents=True, exist_ok=True)
        self.previous_runs_dir = main_module.RUNS_DIR
        self.previous_production_mode = main_module.DEFAULT_PRODUCTION_MODE
        main_module.RUNS_DIR = self.runs_dir
        auth_service._SESSIONS.clear()
        self.client = TestClient(main_module.app)

    def tearDown(self) -> None:
        main_module.RUNS_DIR = self.previous_runs_dir
        main_module.DEFAULT_PRODUCTION_MODE = self.previous_production_mode
        auth_service._SESSIONS.clear()
        self.tempdir.cleanup()

    def _login_admin(self) -> None:
        response = self.client.post(
            "/auth/login",
            json={"actor_name": "contract-test", "role": "admin", "access_code": "spot-local"},
        )
        self.assertEqual(response.status_code, 200, response.text)

    def test_read_endpoints_require_authentication(self) -> None:
        self.assertEqual(self.client.get("/runs").status_code, 401)
        self.assertEqual(self.client.get("/uploads").status_code, 401)
        self.assertEqual(self.client.get("/classify/status/example-run").status_code, 401)

    def test_review_update_rejects_nonexistent_row(self) -> None:
        self._login_admin()
        run_id = f"contract-run-{uuid.uuid4().hex[:8]}"
        _prepare_synthetic_run(self.runs_dir, run_id)
        response = self.client.post(
            f"/runs/{run_id}/review-rows/999",
            json={"review_state": "reviewed", "review_decision": "confirm", "reviewer_note": "should fail"},
        )
        self.assertEqual(response.status_code, 404, response.text)

    def test_signoff_requires_review_completion(self) -> None:
        self._login_admin()
        run_id = f"contract-run-{uuid.uuid4().hex[:8]}"
        _prepare_synthetic_run(self.runs_dir, run_id)
        blocked = self.client.post(
            f"/runs/{run_id}/signoff",
            json={"decision": "accepted_with_conditions", "note": "blocked until review is complete"},
        )
        self.assertEqual(blocked.status_code, 409, blocked.text)

        reviewed = self.client.post(
            f"/runs/{run_id}/review-rows/2",
            json={"review_state": "reviewed", "review_decision": "confirm", "reviewer_note": "done"},
        )
        self.assertEqual(reviewed.status_code, 200, reviewed.text)
        stored_row = fetch_run_row(runs_dir=self.runs_dir, run_id=run_id, row_index=2)
        self.assertIsNotNone(stored_row)
        self.assertEqual((stored_row or {}).get("review_state"), "reviewed")
        self.assertEqual((stored_row or {}).get("review_decision"), "confirm")

        allowed = self.client.post(
            f"/runs/{run_id}/signoff",
            json={"decision": "accepted_with_conditions", "note": "all review rows complete"},
        )
        self.assertEqual(allowed.status_code, 200, allowed.text)

    def test_archive_existing_run_state_preserves_old_attempt(self) -> None:
        run_id = "archive-check"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        marker = run_path / "marker.txt"
        marker.write_text("keep me", encoding="utf-8")

        main_module._archive_existing_run_state(run_id)

        self.assertFalse(run_path.exists())
        history_root = self.runs_dir / "_history"
        archived = list(history_root.glob(f"{run_id}-*"))
        self.assertEqual(len(archived), 1)
        self.assertEqual((archived[0] / "marker.txt").read_text(encoding="utf-8"), "keep me")

    def test_intake_writes_row_manifest_for_accepted_upload(self) -> None:
        upload_id = "manifest-upload"
        record = intake_workbook(
            runs_dir=self.runs_dir,
            ssot_path=PROJECT_ROOT / "ssot/ssot.json",
            upload_id=upload_id,
            filename="manifest.xlsx",
            content=(PROJECT_ROOT / "samples" / "sample_germany.xlsx").read_bytes(),
        )
        self.assertEqual(record["status"], "accepted")
        manifest_path = Path(record["validation"]["row_manifest_path"])
        self.assertTrue(manifest_path.exists())
        lines = [json.loads(line) for line in manifest_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        self.assertGreaterEqual(len(lines), 1)
        self.assertEqual(lines[0]["sequence_index"], 1)
        self.assertEqual(lines[0]["row_index"], 2)
        self.assertTrue(str(lines[0]["item_number"]).strip())
        self.assertTrue(lines[0]["row_hash"])
        self.assertTrue(lines[0]["post_text_sha256"])
        stored_rows = fetch_upload_rows_for_segment(runs_dir=self.runs_dir, upload_id=upload_id, row_start=1, row_end=1)
        self.assertEqual(len(stored_rows), 1)
        self.assertEqual(stored_rows[0]["row_index"], 2)
        self.assertEqual(stored_rows[0]["item_number"], lines[0]["item_number"])
        self.assertEqual(stored_rows[0]["row_hash"], lines[0]["row_hash"])

    def test_segment_input_workbook_from_entries_returns_row_manifest_entries(self) -> None:
        segment_path = self.runs_dir / "segment-input.xlsx"
        manifest_entries = build_segment_input_workbook_from_entries(
            segment_path,
            [
                {
                    "sequence_index": 1,
                    "row_index": 2,
                    "item_number": "1",
                    "post_text": "Synthetic contract test row",
                    "source_category": "",
                    "row_hash": "hash-1",
                    "post_text_sha256": "sha-1",
                    "post_text_length": 27,
                }
            ],
        )
        self.assertTrue(segment_path.exists())
        self.assertEqual(len(manifest_entries), 1)
        self.assertEqual(manifest_entries[0]["sequence_index"], 1)
        self.assertEqual(manifest_entries[0]["row_index"], 2)
        self.assertEqual(manifest_entries[0]["item_number"], "1")
        self.assertEqual(manifest_entries[0]["row_hash"], "hash-1")

    def test_ensure_upload_rows_materialized_backfills_legacy_upload(self) -> None:
        upload_id = "legacy-upload"
        upload_dir = self.runs_dir / "uploads" / upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = upload_dir / "legacy.xlsx"
        workbook_path.write_bytes((PROJECT_ROOT / "samples" / "sample_germany.xlsx").read_bytes())
        expected_count = len(read_input_rows(workbook_path, load_ssot(PROJECT_ROOT / "ssot/ssot.json")))
        upload_record = {
            "upload_id": upload_id,
            "filename": "legacy.xlsx",
            "stored_path": str(workbook_path),
            "bytes": workbook_path.stat().st_size,
            "status": "accepted",
            "created_at": 1771000000,
            "validation": {"accepted": True, "row_count": expected_count},
        }
        from backend.services.ops_db_service import record_upload

        (upload_dir / "upload.json").write_text(json.dumps(upload_record), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=upload_record)
        self.assertEqual(
            fetch_upload_rows_for_segment(runs_dir=self.runs_dir, upload_id=upload_id, row_start=1, row_end=5),
            [],
        )

        count = ensure_upload_rows_materialized(
            runs_dir=self.runs_dir,
            ssot_path=PROJECT_ROOT / "ssot/ssot.json",
            upload_id=upload_id,
        )

        self.assertEqual(count, expected_count)
        stored_rows = fetch_upload_rows_for_segment(runs_dir=self.runs_dir, upload_id=upload_id, row_start=1, row_end=5)
        self.assertEqual(len(stored_rows), 5)
        self.assertEqual(stored_rows[0]["sequence_index"], 1)
        self.assertEqual(stored_rows[0]["row_index"], 2)
        manifest_path = upload_dir / "row_manifest.jsonl"
        self.assertTrue(manifest_path.exists())
        manifest_lines = [json.loads(line) for line in manifest_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        self.assertEqual(len(manifest_lines), expected_count)
        self.assertEqual(manifest_lines[0]["row_index"], 2)

    def test_production_payload_restrictions_block_arbitrary_paths(self) -> None:
        main_module.DEFAULT_PRODUCTION_MODE = True

        main_module._assert_allowed_classify_payload({"upload_id": "upload-1", "language": "de", "review_mode": "partial", "limit": 1})

        with self.assertRaises(HTTPException):
            main_module._assert_allowed_classify_payload({"input": "/tmp/file.xlsx"})

        with self.assertRaises(HTTPException):
            main_module._assert_allowed_classify_payload({"output": "/tmp/out.xlsx"})

    def test_recover_restarts_interrupted_segment_worker_run(self) -> None:
        self._login_admin()
        upload_id = "recover-upload"
        upload_dir = self.runs_dir / "uploads" / upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = upload_dir / "recover.xlsx"
        workbook_path.write_bytes(_build_workbook_bytes())
        upload_record = {
            "upload_id": upload_id,
            "filename": "recover.xlsx",
            "stored_path": str(workbook_path),
            "bytes": workbook_path.stat().st_size,
            "status": "accepted",
            "created_at": 1771000000,
            "validation": {"accepted": True, "row_count": 1},
        }
        from backend.services.ops_db_service import record_upload, register_run

        record_upload(runs_dir=self.runs_dir, record=upload_record)
        run_id = "recover-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(workbook_path),
            output_path=str(run_path / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "PROCESSING", "progress_percentage": 0.0, "processed_rows": 0, "total_rows": 1}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": 999999, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )
        with patch.object(
            main_module,
            "_start_classify_run",
            return_value={"status": "restarted", "run_id": run_id, "pid": 12345},
        ) as mocked_start:
            response = self.client.post(f"/runs/{run_id}/recover")
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "restarted")
        mocked_start.assert_called_once()

    def test_heal_restarts_stalled_segment_worker_run(self) -> None:
        self._login_admin()
        upload_id = "heal-upload"
        upload_dir = self.runs_dir / "uploads" / upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = upload_dir / "heal.xlsx"
        workbook_path.write_bytes(_build_workbook_bytes())
        run_id = "heal-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(workbook_path),
            output_path=str(run_path / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "PROCESSING", "progress_percentage": 5.0, "processed_rows": 5, "total_rows": 100}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": 54321, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )
        with patch.object(main_module, "_resolve_run_process_pid", return_value=54321), patch.object(
            main_module, "_signal_run_process"
        ) as mocked_signal, patch.object(
            main_module, "_wait_for_pid_exit", return_value=True
        ), patch.object(
            main_module, "_start_classify_run", return_value={"status": "restarted", "run_id": run_id, "pid": 12345}
        ) as mocked_start:
            response = self.client.post(f"/runs/{run_id}/heal")
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "restarted")
        mocked_signal.assert_called_once()
        mocked_start.assert_called_once()

    def test_native_runtime_suspend_marks_active_run_for_resume(self) -> None:
        self._login_admin()
        run_id = "native-suspend-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path="samples/sample_germany.xlsx",
            output_path=str(run_path / "output.xlsx"),
            upload_id="upload-1",
            language="de",
            review_mode="partial",
            start_payload={"upload_id": "upload-1", "language": "de", "review_mode": "partial"},
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "PROCESSING", "progress_percentage": 50.0, "processed_rows": 5, "total_rows": 10}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": 43210, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )
        with patch.object(main_module, "_pid_alive", side_effect=lambda pid: int(pid or 0) == 43210), patch.object(
            main_module,
            "_pid_command",
            return_value=f"python backend/segment_worker.py --run-id {run_id}",
        ), patch.object(
            run_state_service,
            "_pid_alive",
            side_effect=lambda pid: int(pid or 0) == 43210,
        ), patch.object(
            run_state_service,
            "_pid_command",
            return_value=f"python backend/segment_worker.py --run-id {run_id}",
        ), patch.object(main_module, "_signal_run_process") as mocked_signal:
            response = self.client.post("/native/runtime/suspend")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], "suspend_requested")
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["active_runs"][0]["run_id"], run_id)
        mocked_signal.assert_called_once()
        control = json.loads((run_path / "control.json").read_text(encoding="utf-8"))
        self.assertTrue(control["shutdown_requested"])
        self.assertEqual(control["shutdown_mode"], "suspend")
        self.assertFalse(control["paused"])

    def test_classify_status_does_not_trust_reused_unrelated_pid(self) -> None:
        self._login_admin()
        run_id = "status-pid-reuse"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path="samples/sample_germany.xlsx",
            output_path=str(run_path / "output.xlsx"),
            upload_id="upload-1",
            language="de",
            review_mode="partial",
            start_payload={"upload_id": "upload-1", "language": "de", "review_mode": "partial"},
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "INTERRUPTED", "processed_rows": 10, "total_rows": 100, "progress_percentage": 10.0}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": 424242, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )
        with patch.object(main_module, "_pid_alive", return_value=True), patch.object(
            main_module,
            "_pid_command",
            return_value="python unrelated_service.py --port 9999",
        ), patch.object(main_module, "_discover_classify_pid", return_value=None):
            response = self.client.get(f"/classify/status/{run_id}")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertFalse(payload["running"])
        self.assertIsNone(payload["pid"])

    def test_retry_upload_run_resumes_existing_state(self) -> None:
        self._login_admin()
        run_id = "retry-upload-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path="samples/sample_germany.xlsx",
            output_path=str(run_path / "output.xlsx"),
            upload_id="upload-1",
            language="de",
            review_mode="partial",
            start_payload={"upload_id": "upload-1", "language": "de", "review_mode": "partial"},
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "FAILED", "processed_rows": 10, "total_rows": 100, "progress_percentage": 10.0}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": None, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )
        with patch.object(
            main_module,
            "_start_classify_run",
            return_value={"status": "restarted", "run_id": run_id, "pid": 12345},
        ) as mocked_start:
            response = self.client.post(f"/runs/{run_id}/retry")
        self.assertEqual(response.status_code, 200, response.text)
        mocked_start.assert_called_once()
        self.assertTrue(bool(mocked_start.call_args.kwargs.get("resume_existing")))

    def test_recovery_candidates_flag_missing_source_as_broken(self) -> None:
        self._login_admin()
        upload_id = "broken-upload"
        upload_dir = self.runs_dir / "uploads" / upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = upload_dir / "missing.xlsx"
        upload_record = {
            "upload_id": upload_id,
            "filename": "missing.xlsx",
            "stored_path": str(workbook_path),
            "bytes": 0,
            "status": "accepted",
            "created_at": 1771000000,
            "validation": {"accepted": True, "row_count": 1},
        }
        from backend.services.ops_db_service import record_upload, register_run

        (upload_dir / "upload.json").write_text(json.dumps(upload_record), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=upload_record)
        run_id = "broken-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(workbook_path),
            output_path=str(run_path / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            state="PROCESSING",
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "PROCESSING", "processed_rows": 1, "total_rows": 10, "progress_percentage": 10.0}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": None, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )
        (run_path / "logs.txt").write_text("[FAILED] missing source workbook\n", encoding="utf-8")

        response = self.client.get("/runs/recovery-candidates/list")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(len(payload["runs"]), 1)
        candidate = payload["runs"][0]
        self.assertEqual(candidate["run_id"], run_id)
        self.assertTrue(candidate["is_broken"])
        self.assertFalse(candidate["can_continue"])
        self.assertIn("managed source workbook is missing from local storage", candidate["broken_reasons"])
        self.assertEqual(candidate["last_log_line"], "[FAILED] missing source workbook")

    def test_delete_run_purges_run_and_managed_source_when_inactive(self) -> None:
        self._login_admin()
        upload_id = "purge-upload"
        upload_dir = self.runs_dir / "uploads" / upload_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = upload_dir / "purge.xlsx"
        workbook_path.write_bytes(_build_workbook_bytes())
        upload_record = {
            "upload_id": upload_id,
            "filename": "purge.xlsx",
            "stored_path": str(workbook_path),
            "bytes": workbook_path.stat().st_size,
            "status": "accepted",
            "created_at": 1771000000,
            "validation": {"accepted": True, "row_count": 1},
        }
        from backend.services.ops_db_service import record_upload, register_run

        (upload_dir / "upload.json").write_text(json.dumps(upload_record), encoding="utf-8")
        record_upload(runs_dir=self.runs_dir, record=upload_record)
        run_id = "purge-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(workbook_path),
            output_path=str(run_path / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        register_run(
            runs_dir=self.runs_dir,
            run_id=run_id,
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            state="FAILED",
        )
        (run_path / "progress.json").write_text(
            json.dumps({"run_id": run_id, "state": "FAILED", "processed_rows": 1, "total_rows": 10, "progress_percentage": 10.0}),
            encoding="utf-8",
        )
        (run_path / "control.json").write_text(
            json.dumps({"run_id": run_id, "pid": None, "paused": False, "output": str(run_path / "output.xlsx")}),
            encoding="utf-8",
        )

        response = self.client.delete(f"/runs/{run_id}?purge_source=true")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], "deleted")
        self.assertFalse(run_path.exists())
        self.assertFalse(upload_dir.exists())
        self.assertIsNone(run_state_service.read_run_record(runs_dir=self.runs_dir, run_id=run_id))

    def test_run_classification_resumes_from_checkpoint(self) -> None:
        workbook_path = self.runs_dir / "resume-input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Item number", "Post text", "Category"])
        ws.append(["1", "First resume row", ""])
        ws.append(["2", "Second resume row", ""])
        wb.save(workbook_path)
        wb.close()

        run_id = "resume-pipeline-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        checkpoint_path = run_dir / "result_checkpoint.jsonl"
        first_hash = stable_row_hash("1", "First resume row")
        first_result = ClassificationResult(
            row_index=2,
            raw_category="Not Antisemitic",
            category="Not Antisemitic",
            confidence=0.91,
            explanation="Checkpointed result.",
            flags=[],
            resolved_model_version="synthetic-model",
        )
        _append_result_checkpoint(checkpoint_path, result_index=0, row_hash=first_hash, result=first_result)

        def _fake_classify_batch(rows, ssot, max_workers, review_mode, model_name="", progress_callback=None, row_completion_callback=None, progress_every=100):
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].row_index, 3)
            second_hash = stable_row_hash(rows[0].item_number, rows[0].post_text)
            second_result = ClassificationResult(
                row_index=3,
                raw_category="Anti-Israel",
                category="Anti-Israel",
                confidence=0.88,
                explanation="Resumed row result.",
                flags=["REVIEW_REQUIRED"],
                resolved_model_version="synthetic-model",
            )
            if row_completion_callback:
                row_completion_callback(1, 1, 0, second_result, second_hash)
            return [second_result], [second_hash]

        with patch("src.pipeline.classify_batch", side_effect=_fake_classify_batch):
            run_classification(
                input_path=workbook_path,
                output_path=run_dir / "output.xlsx",
                run_id=run_id,
                run_language="de",
                review_mode="partial",
                ssot_path=PROJECT_ROOT / "ssot/ssot.json",
                runs_dir=self.runs_dir,
                max_workers=1,
                progress_every=1,
                resume_existing=True,
            )

        progress = json.loads((run_dir / "progress.json").read_text(encoding="utf-8"))
        self.assertEqual(progress["state"], "COMPLETED")
        self.assertEqual(progress["total_rows"], 2)
        checkpoint_lines = [json.loads(line) for line in checkpoint_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        self.assertEqual(len(checkpoint_lines), 2)

    def test_run_classification_persists_canonical_rows_during_execution(self) -> None:
        workbook_path = self.runs_dir / "canonical-input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Item number", "Post text", "Category"])
        ws.append(["1", "Canonical execution row", ""])
        wb.save(workbook_path)
        wb.close()

        run_id = "canonical-runtime-run"
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(workbook_path),
            output_path=str(run_dir / "output.xlsx"),
            upload_id=None,
            language="de",
            review_mode="partial",
            start_payload={"language": "de", "review_mode": "partial"},
        )
        attempt = start_run_attempt(runs_dir=self.runs_dir, run_id=run_id, attempt_type="start")

        def _fake_classify_batch(rows, ssot, max_workers, review_mode, model_name="", progress_callback=None, row_completion_callback=None, progress_every=100):
            result = ClassificationResult(
                row_index=2,
                raw_category="Anti-Israel",
                category="Anti-Israel",
                confidence=0.82,
                explanation="Canonical runtime persistence.",
                flags=["REVIEW_REQUIRED"],
                resolved_model_version="synthetic-model",
            )
            row_hash = stable_row_hash(rows[0].item_number, rows[0].post_text)
            if row_completion_callback:
                row_completion_callback(1, 1, 0, result, row_hash)
            return [result], [row_hash]

        with patch("src.pipeline.classify_batch", side_effect=_fake_classify_batch):
            run_classification(
                input_path=workbook_path,
                output_path=run_dir / "output.xlsx",
                run_id=run_id,
                run_language="de",
                review_mode="partial",
                ssot_path=PROJECT_ROOT / "ssot/ssot.json",
                runs_dir=self.runs_dir,
                max_workers=1,
                progress_every=1,
                canonical_runs_dir=self.runs_dir,
                canonical_run_id=run_id,
                canonical_attempt_id=str(attempt["attempt_id"]),
            )

        stored_row = fetch_run_row(runs_dir=self.runs_dir, run_id=run_id, row_index=2)
        self.assertIsNotNone(stored_row)
        self.assertEqual((stored_row or {}).get("assigned_category"), "Anti-Israel")
        self.assertEqual((stored_row or {}).get("attempt_id"), attempt["attempt_id"])
        self.assertTrue((stored_row or {}).get("review_required"))

    def test_migrate_row_state_from_output_backfills_canonical_rows(self) -> None:
        self._login_admin()
        run_id = f"migrate-output-{uuid.uuid4().hex[:8]}"
        _prepare_synthetic_run(self.runs_dir, run_id)

        response = self.client.post(f"/runs/{run_id}/migrate-row-state")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertTrue(payload["output_present"])
        self.assertEqual(payload["output_rows_migrated"], 1)

        stored_row = fetch_run_row(runs_dir=self.runs_dir, run_id=run_id, row_index=2)
        self.assertIsNotNone(stored_row)
        self.assertEqual((stored_row or {}).get("assigned_category"), "Anti-Israel")
        self.assertTrue((stored_row or {}).get("review_required"))

    def test_review_queue_does_not_parse_output_without_explicit_migration(self) -> None:
        run_id = f"output-read-bridge-{uuid.uuid4().hex[:8]}"
        _prepare_synthetic_run(self.runs_dir, run_id)

        queue_before = build_review_queue(runs_dir=self.runs_dir, run_id=run_id)
        self.assertIsNotNone(queue_before)
        self.assertEqual(queue_before["rows"], [])
        self.assertIsNone(fetch_run_row(runs_dir=self.runs_dir, run_id=run_id, row_index=2))

        summary = run_state_service.migrate_run_rows_to_canonical(runs_dir=self.runs_dir, run_id=run_id)
        self.assertEqual(summary["output_rows_migrated"], 1)

        queue_after = build_review_queue(runs_dir=self.runs_dir, run_id=run_id)
        self.assertIsNotNone(queue_after)
        self.assertEqual(len(queue_after["rows"]), 1)
        self.assertEqual(queue_after["rows"][0]["row_index"], 2)
        self.assertEqual(queue_after["rows"][0]["assigned_category"], "Anti-Israel")

    def test_migrate_row_state_from_checkpoints_imports_nonreview_rows(self) -> None:
        self._login_admin()
        upload_id = f"checkpoint-upload-{uuid.uuid4().hex[:8]}"
        record = intake_workbook(
            runs_dir=self.runs_dir,
            ssot_path=PROJECT_ROOT / "ssot/ssot.json",
            upload_id=upload_id,
            filename="checkpoint.xlsx",
            content=(PROJECT_ROOT / "samples" / "sample_germany.xlsx").read_bytes(),
        )
        self.assertEqual(record["status"], "accepted")

        run_id = f"checkpoint-run-{uuid.uuid4().hex[:8]}"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(Path(record["stored_path"])),
            output_path=str(run_path / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        checkpoint_dir = run_path / "_segments" / f"{upload_id}-seg-00001" / "runs" / f"{run_id}-segment-00001"
        checkpoint_dir.mkdir(parents=True, exist_ok=True)
        _append_result_checkpoint(
            checkpoint_dir / "result_checkpoint.jsonl",
            result_index=0,
            row_hash=stable_row_hash(
                "1",
                "There are a lot of cultural barbarians in Germany now who hate us as Germans and abuse us as zombies or white bread. We are all pigs. That is racist, isn't it?",
            ),
            result=ClassificationResult(
                row_index=2,
                raw_category="Not Antisemitic",
                category="Not Antisemitic",
                confidence=0.91,
                explanation="Checkpointed non-review row.",
                flags=[],
                resolved_model_version="synthetic-model",
            ),
        )

        response = self.client.post(f"/runs/{run_id}/migrate-row-state?include_output=false")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["checkpoint_files_seen"], 1)
        self.assertEqual(payload["checkpoint_rows_migrated"], 1)
        self.assertFalse(payload["output_present"])

        stored_row = fetch_run_row(runs_dir=self.runs_dir, run_id=run_id, row_index=2)
        self.assertIsNotNone(stored_row)
        self.assertEqual((stored_row or {}).get("assigned_category"), "Not Antisemitic")
        self.assertFalse(bool((stored_row or {}).get("review_required")))

    def test_review_queue_reads_live_checkpoint_rows(self) -> None:
        upload_id = "live-review-upload"
        record = intake_workbook(
            runs_dir=self.runs_dir,
            ssot_path=PROJECT_ROOT / "ssot/ssot.json",
            upload_id=upload_id,
            filename="live-review.xlsx",
            content=(PROJECT_ROOT / "samples" / "sample_germany.xlsx").read_bytes(),
        )
        self.assertEqual(record["status"], "accepted")
        run_id = "live-review-run"
        run_path = self.runs_dir / run_id
        run_path.mkdir(parents=True, exist_ok=True)
        create_run_record(
            runs_dir=self.runs_dir,
            run_id=run_id,
            input_path=str(Path(record["stored_path"])),
            output_path=str(run_path / "output.xlsx"),
            upload_id=upload_id,
            language="de",
            review_mode="partial",
            start_payload={"upload_id": upload_id, "language": "de", "review_mode": "partial"},
        )
        checkpoint_dir = run_path / "_segments" / f"{upload_id}-seg-00001" / "runs" / f"{run_id}-segment-00001"
        checkpoint_dir.mkdir(parents=True, exist_ok=True)
        _append_result_checkpoint(
            checkpoint_dir / "result_checkpoint.jsonl",
            result_index=0,
            row_hash=stable_row_hash("1", "There are a lot of cultural barbarians in Germany now who hate us as Germans and abuse us as zombies or white bread. We are all pigs. That is racist, isn't it?"),
            result=ClassificationResult(
                row_index=2,
                raw_category="Anti-Israel",
                category="Anti-Israel",
                confidence=0.77,
                explanation="Checkpointed flagged row.",
                flags=["REVIEW_REQUIRED"],
                resolved_model_version="synthetic-model",
            ),
        )
        queue = build_review_queue(runs_dir=self.runs_dir, run_id=run_id)
        self.assertIsNotNone(queue)
        rows = queue["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["row_index"], 2)
        self.assertEqual(rows[0]["review_state"], "pending")
        self.assertTrue(rows[0]["post_text"])

    def test_artifact_manifest_catalog_includes_review_and_control_artifacts(self) -> None:
        run_dir = self.runs_dir / "artifact-manifest-run"
        run_dir.mkdir(parents=True, exist_ok=True)
        for name in ["progress.json", "processing_stats.json", "review_state.json", "signoff.json", "action_log.jsonl", "control.json"]:
            (run_dir / name).write_text("{}", encoding="utf-8")
        manifest = build_artifact_manifest(run_dir=run_dir, sha256_file=lambda path: f"sha-{path.name}")
        artifacts = manifest["artifacts"]
        self.assertIn("review_state.json", artifacts)
        self.assertIn("signoff.json", artifacts)
        self.assertIn("action_log.jsonl", artifacts)
        self.assertIn("control.json", artifacts)
        self.assertIn("processing_stats.json", artifacts)


if __name__ == "__main__":
    suite = unittest.defaultTestLoader.loadTestsFromTestCase(BackendContractRegressionTests)
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    raise SystemExit(0 if result.wasSuccessful() else 1)
